#!/usr/bin/env python
# coding: utf-8

# In[56]:


import pandas as pd
import numpy as np
from pandas.tseries.offsets import MonthEnd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import xlwings as xw
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.cell import get_column_letter
import getpass
import sys
from copy import copy
from spire.xls.common import *
from spire.xls import *
import warnings
warnings.filterwarnings("ignore")
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList

#Functions

def read_raw_data(filepath):
    df = pd.read_excel(filepath, index_col=None, header=None)
    return(df)

def reorder_column(df,columnname,position):
    df.insert(position,columnname,df.pop(columnname))
    return df

def concatenate_dfs(dfs):
    df = pd.concat(dfs, ignore_index=True)
    return df

def read_ncr(path,month):
    df = pd.read_excel(path,header=None)
    row_no, col_no = coordinate_finder_insheet_df(df,"Quantity",1)
    row_no1, col_no = coordinate_finder_insheet_df(df,"Report Totals",1)
    df.columns = df.iloc[row_no,:]
    df = df.iloc[row_no+1:row_no1,:]
    df = df.dropna(axis=1,how="all")
    df["Month"] = month
    df = reorder_column(df,"Month",1)
    df.columns=["Facility","Month","Quantity","Charges","Fee Schedule Allowed","Total Paid","Patient Paid","Insurance Paid","Patient Adjustment","Insurance Adjustments","Remaining Balance","Net Collection Rate"]
    return df

def read_fin(path,month):
    df = pd.read_excel(path,header=None)
    row_no, col_no = coordinate_finder_insheet_df(df,"Charge Quantity",1)
    row_no1, col_no = coordinate_finder_insheet_df(df,"Report Totals",1)
    df.columns = df.iloc[row_no,:]
    df = df.iloc[row_no+1:row_no1,:]
    df = df.dropna(axis=1,how="all")
    df["Month"] = month
    df = reorder_column(df,"Month",1)
    df.columns = ["Facility","Month","Charge Quantity","Charges","Total Adjustments","Credit Memos","Patient Payments","Insurance Payments","Total Payments","Refunds","Patient Applied Payments","Insurance Applied Payments","Applied Interest","Total Applied Payments"]
    return df

def read_ar(path,month):
    df = pd.read_excel(path,header=None)
    row_no, col_no = coordinate_finder_insheet_df(df,"Current",1)
    row_no1, col_no = coordinate_finder_insheet_df(df,"Report Totals",1)
    df.columns = df.iloc[row_no,:]
    df = df.iloc[row_no+1:row_no1,:]
    df = df.dropna(axis=1,how="all")
    df["Month"] = month
    df = reorder_column(df,"Month",1)
    df.columns = ["Facility","Snapshot Month","0-30","31-60","61-90","91-120","121-150","150+","Totals"]
    return df

def read_ar_cat(path,month,facility):
    df = pd.read_excel(path,header=None)
    row_no, col_no = coordinate_finder_insheet_df(df,"0-30",1)
    df.columns = df.iloc[row_no,:]
    df = df.iloc[row_no+1:,col_no:]
    df = df.dropna(axis=1,how="all")
    df["Month"] = month
    df = reorder_column(df,"Month",0)
    df["Facility"] = facility
    df = reorder_column(df,"Facility",0)
    df.columns = ["Facility","Month","0-30","31-60","61-90","91-120","120+","Total"]
    df = list(df.iloc[-3,:])
    reshaped_values = [df[i:i+len(column_names)] for i in range(0, len(df), len(column_names))]
    df = pd.DataFrame(reshaped_values, columns=column_names)
    return df

def read_year_fin(path,month,facility):
    df = pd.read_excel(path,header=None)
    row_no, col_no = coordinate_finder_insheet_df(df,"Charges",1)
    row_no1, col_no = coordinate_finder_insheet_df(df,month.month,1)
    df.columns = df.iloc[row_no,:]
    df = df.iloc[row_no1:row_no1+1,col_no+1:]
    df = df.dropna(axis=1,how="all")
    df["Month"] = month
    df = reorder_column(df,"Month",0)
    df.columns = ['Month','Charges','Cash','Checks','Credit Crards', 'Total Payments','Refunds','Net Payments','Adjustments','Prepayments','Prepayments Applied','Net Change in Patient Balance'] 
    df["Facility"] = facility
    df = reorder_column(df,"Facility",0)
    return df

def read_collection(path,month,facility):
    df = pd.read_excel(path,header=None)
    row_no, col_no = coordinate_finder_insheet_df(df,"Total Charges",1)
    row_no1, col_no = coordinate_finder_insheet_df(df,"Report Totals:",1)
    df.columns = df.iloc[row_no,:]
    df = df.iloc[row_no1:,col_no+1:]
    df = df.dropna(axis=1,how="all")
    df["Month"] = month
    df = reorder_column(df,"Month",0)
    df.columns = ["Month","Total Charges","Total Payments","% of Payments","Total Adjustments","% of Adjustments"]
    df["Facility"] = facility
    df = reorder_column(df,"Facility",0)
    return df

#Here we find a particular values coordinate in a sheet with the recurrence at what time
def coordinate_finder_insheet_df(df, valuetofind, recurrence):   
    found = False
    count = 0
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if df.iloc[i, j] == valuetofind:
                row_no = i
                col_no = j
                count += 1
                if count == recurrence:
                    found = True
                    break
        if found:
            break
    if found:
        return (row_no, col_no)
    else:
        return (0, 0) 

def copy_paste_raw(Template,sheetname,df,from_row,from_col,to_col):
    
    Sheet=Template[sheetname]

    row_number=from_row
    i=0
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row-1+len(df),min_col=from_col,max_col=to_col):
        j=0
        col_number=from_col
        for cell in row:
            Sheet.cell(row=row_number,column=col_number).value=df.iloc[i,j]
            j+=1
            col_number+=1
        i+=1
        row_number+=1 
    return(Template)

def format_cells(Template,sheetname,row_no_start,row_no_end,col_no_start,col_no_end):
    sheet=Template[sheetname]
    for row in range(row_no_start, row_no_end+1):
        for col_no in range(col_no_start,col_no_end+1):
            cell = sheet.cell(row=row, column=col_no)
            cell.border = copy(sheet.cell(row=4, column=col_no).border)
            cell.fill = copy(sheet.cell(row=4, column=col_no).fill)
            cell.number_format = copy(sheet.cell(row=4, column=col_no).number_format)
            cell.font = copy(sheet.cell(row=4, column=col_no).font)
            cell.alignment = copy(sheet.cell(row=4, column=col_no).alignment)
        
    return Template

def bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_right_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_border_one_cell(Template,sheetname,row_no,col_no):
    sheet=Template[sheetname]
    border = Border(left=Side(style='medium'), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))
    
    cell = sheet.cell(row=row_no, column=col_no)
    cell.border = border
        
    return Template

def right_border_one_cell(Template,sheetname,row_no,col_no):
    sheet=Template[sheetname]
    border = Border(left=Side(style=None), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))
    
    cell = sheet.cell(row=row_no, column=col_no)
    cell.border = border
        
    return Template

def left_right_border(Template,sheetname,row_start,row_end,col_no):
    sheet=Template[sheetname]
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for row in range(row_start, row_end+1):
        cell = sheet.cell(row=row, column=col_no)
        cell.border = border
        
    return Template

#Triggers calculation in the sheet
def perform_calc(filepath):
    wb = xw.Book(filepath)
    app = xw.App(visible=False)
    wb.app.calculate()
    wb.save()
    wb.close()
    app.kill()
    
def extend_formulas(Template,sheetname,df,from_row,from_col,to_col):
    Sheet=Template[sheetname]
    i=from_row
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row+len(df)-1,min_col=from_col,max_col=to_col):
        j=from_col
        for cell in row:
            Sheet.cell(row=i,column=j).value=Translator(Sheet.cell(row=from_row-1,column=j).value,origin=Sheet.cell(row=from_row-1,column=j).coordinate).translate_formula(Sheet.cell(row=i,column=j).coordinate)    
            j+=1
        i+=1
    return(Template)


import time
start_time = time.time()

# Get the current system's username
username = getpass.getuser()
print("Username:", username)

#New user need to store this
path_folder = r"C:\Users\pragna_kandagatla\Desktop\Automation\SLA Tracker"

#Checking the files required i.e setup files required and storing
for filename in os.listdir(path_folder):
    file_path4 = os.path.join(path_folder, filename)
    if filename.endswith("Tucson_SLAs.xlsx"):
        template_path = file_path4

folder_path_nt =input("Paste the path of Nextech raw data folder= ")
folder_path_nc =input("Paste the path of NexCloud raw data folder= ")
current_month = pd.to_datetime(input("Please enter the start date of the month in (mm/dd/yyyy) format = "))
print("\n")
folder_path_nt=folder_path_nt.replace("\\","/").replace("\"","")
folder_path_nc=folder_path_nc.replace("\\","/").replace("\"","")

file_saving_location = f"C:/Users/{username}/Downloads/Tucson_SLAs.xlsx"

ncr_path = ""
ar_path = ""
fin_path = ""

Template=load_workbook(filename=template_path)

#storing all the sheets as a dataframe
template_ar_summary=pd.read_excel(template_path,sheet_name="Accounts Receivable Summary",header=None)
template_fin_summary=pd.read_excel(template_path,sheet_name="Financial Activity Report",header=None)
template_charges_summary=pd.read_excel(template_path,sheet_name="Charges Collected Summary",header=None)
template_ar_by_ins=pd.read_excel(template_path,sheet_name="AR by Insurance",header=None)
template_month_fin_summary=pd.read_excel(template_path,sheet_name="Monthly Fin Summary",header=None)
template_collection_rate=pd.read_excel(template_path,sheet_name="Collection Rate by Category",header=None)
    
for j, filename in enumerate(os.listdir(folder_path_nt)):
    file_path = os.path.join(folder_path_nt, filename)
    for i, filename2 in enumerate(os.listdir(file_path)):
        file_path2 = os.path.join(file_path, filename2)
        file_path2 = file_path2.replace("\\","/").replace("\"","")
        read_df = read_raw_data(file_path2)
        row_no_ncr, col_ncr = coordinate_finder_insheet_df(read_df, "Charges With Collected Amount Summary", 1)
        row_no_ar, col_ar = coordinate_finder_insheet_df(read_df, "Accounts Receivable Summary", 1)
        row_no_fin, col_fin = coordinate_finder_insheet_df(read_df, "Financial Activity Summary", 1)
        if read_df.iloc[row_no_ncr+1, col_ncr] == "Effective Date:":
            ncr_path = file_path
        elif read_df.iloc[row_no_ar+1, col_ar] == "Transaction Date:":
            ar_path = file_path
        elif read_df.iloc[row_no_fin+1, col_fin] == "Practice:":
            fin_path = file_path
print('Nextech NCR Path: ',ncr_path)
print('Nextech AR Path: ',ar_path)
print('Nextech Fin Path: ',fin_path)


#--------------------------------------------------Execution Starts-------------------------------------

#----------------------------------- Reading data----------------------------------------------------------
#--------------------------Nextech-----------------------
ar_df = pd.DataFrame(columns=["Facility","Snapshot Month","0-30","31-60","61-90","91-120","121-150","150+","Totals"])
ncr_df = pd.DataFrame(columns=["Facility","Month","Quantity","Charges","Fee Schedule Allowed","Total Paid","Patient Paid","Insurance Paid","Patient Adjustment","Insurance Adjustments","Remaining Balance","Net Collection Rate"])
fin_df = pd.DataFrame(columns=["Facility","Month","Charge Quantity","Charges","Total Adjustments","Credit Memos","Patient Payments","Insurance Payments","Total Payments","Refunds","Patient Applied Payments","Insurance Applied Payments","Applied Interest","Total Applied Payments"])

for j, filename in enumerate(os.listdir(ar_path)):
    file_path = os.path.join(ar_path, filename)
    file_path=file_path.replace("\\","/").replace("\"","")
    df = read_ar(file_path,current_month)
    ar_df = concatenate_dfs([ar_df,df])

for j, filename in enumerate(os.listdir(fin_path)):
    file_path = os.path.join(fin_path, filename)
    file_path=file_path.replace("\\","/").replace("\"","")
    df = read_fin(file_path,current_month)
    fin_df = concatenate_dfs([fin_df,df])

for j, filename in enumerate(os.listdir(ncr_path)):
    file_path = os.path.join(ncr_path, filename)
    file_path=file_path.replace("\\","/").replace("\"","")
    df = read_ncr(file_path,current_month)
    ncr_df = concatenate_dfs([ncr_df,df])

print("Raw Data Files for Nextech are loaded")

#------------------------------NexCloud-----------------------------------------------
ar_nc_df = pd.DataFrame(columns=["Facility","Month","0-30","31-60","61-90","91-120","120+","Total"])
collection_df = pd.DataFrame(columns=["Facility","Month","Total Charges","Total Payments","% of Payments","Total Adjustments","% of Adjustments"])
year_fin_df = pd.DataFrame(columns=['Facility','Month','Charges','Cash','Checks','Credit Crards', 'Total Payments','Refunds','Net Payments','Adjustments','Prepayments','Prepayments Applied','Net Change in Patient Balance'])

for j, filename in enumerate(os.listdir(folder_path_nc)):
    file_path = os.path.join(folder_path_nc, filename)
    for i, filename2 in enumerate(os.listdir(file_path)):
        file_path2 = os.path.join(file_path, filename2)
        file_path2 = file_path2.replace("\\","/").replace("\"","")
        read_df = read_raw_data(file_path2)
        row_no_ncr, col_ncr = coordinate_finder_insheet_df(read_df, "Total Payments", 1)
        row_no_ar, col_ar = coordinate_finder_insheet_df(read_df, "Date billed", 1)
        row_no_fin, col_fin = coordinate_finder_insheet_df(read_df, "Month", 1)
        if read_df.iloc[row_no_ncr, col_ncr+1] == "% of Payments":
            ncr_path = file_path2
        elif read_df.iloc[row_no_ar, col_ar+2] == "Description":
            ar_path = file_path2
        elif read_df.iloc[row_no_fin, col_fin+2] == "Charges":
            fin_path = file_path2
    ar_nc_df = concatenate_dfs([ar_nc_df,read_ar_cat(ar_path,current_month,filename)])
    collection_df = concatenate_dfs([collection_df,read_collection(ncr_path,current_month,filename)])
    year_fin_df =  concatenate_dfs([year_fin_df,read_year_fin(fin_path,current_month,filename)]) 
    
print("Raw Data Files for NexCloud are loaded")
    
#------------------------------------Pasting Raw Data and formatting in Excel-----------------------------------

#----------------------AR Summary-------------------------------
row_no,col_no=coordinate_finder_insheet_df(template_ar_summary,"Facility",1)
to_row,to_col=coordinate_finder_insheet_df(template_ar_summary,"Totals",1)
to_row,col_no1=coordinate_finder_insheet_df(template_ar_summary,"90+ AR",1)

#Paste values in excel
Template = copy_paste_raw(Template,"Accounts Receivable Summary",ar_df,len(template_ar_summary)+1,col_no+1,to_col+1)
Template = extend_formulas(Template,"Accounts Receivable Summary",ar_df,len(template_ar_summary)+1,col_no1+1,col_no1+1)
Template = format_cells(Template,"Accounts Receivable Summary",len(template_ar_summary),len(template_ar_summary)+len(ar_df),col_no+1,col_no1+1)
Template = bottom_border(Template,"Accounts Receivable Summary", len(template_ar_summary)+len(ar_df),col_no+1,col_no1+1)
Template = left_border_one_cell(Template,"Accounts Receivable Summary", len(template_ar_summary)+len(ar_df),col_no+1)
Template = right_border_one_cell(Template,"Accounts Receivable Summary", len(template_ar_summary)+len(ar_df),col_no1+1)


#---------------------Financial Activity Report-------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet_df(template_fin_summary,"Facility",1)
to_row,to_col=coordinate_finder_insheet_df(template_fin_summary,"Total Applied Payments",1)

#Paste values in excel
Template = copy_paste_raw(Template,"Financial Activity Report",fin_df,len(template_fin_summary)+1,col_no+1,to_col+1)
Template = format_cells(Template,"Financial Activity Report",len(template_fin_summary),len(template_fin_summary)+len(fin_df),col_no+1,to_col+1)
Template = bottom_border(Template,"Financial Activity Report", len(template_fin_summary)+len(fin_df),col_no+1,to_col+1)
Template = left_border_one_cell(Template,"Financial Activity Report", len(template_fin_summary)+len(fin_df),col_no+1)
Template = right_border_one_cell(Template,"Financial Activity Report", len(template_fin_summary)+len(fin_df),to_col+1)



#---------------------Charges Collected Summary Report-------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet_df(template_charges_summary,"Facility",1)
to_row,to_col=coordinate_finder_insheet_df(template_charges_summary,"Net Collection Rate",1)

#Paste values in excel
Template = copy_paste_raw(Template,"Charges Collected Summary",ncr_df,len(template_charges_summary)+1,col_no+1,to_col+1)
Template = format_cells(Template,"Charges Collected Summary",len(template_charges_summary),len(template_charges_summary)+len(ncr_df),col_no+1,to_col+1)
Template = bottom_border(Template,"Charges Collected Summary", len(template_charges_summary)+len(ncr_df),col_no+1,to_col+1)
Template = left_border_one_cell(Template,"Charges Collected Summary", len(template_charges_summary)+len(ncr_df),col_no+1)
Template = right_border_one_cell(Template,"Charges Collected Summary", len(template_charges_summary)+len(ncr_df),to_col+1)

#----------------------AR Insurance--------------------------------------
row_no,col_no=coordinate_finder_insheet_df(template_ar_by_ins,"Facility",1)
to_row,to_col=coordinate_finder_insheet_df(template_ar_by_ins,"Totals",1)
to_row,col_no1=coordinate_finder_insheet_df(template_ar_by_ins,"90+ AR",1)

#Paste values in excel
Template = copy_paste_raw(Template,"AR by Insurance",ar_nc_df,len(template_ar_by_ins)+1,col_no+1,to_col+1)
Template = extend_formulas(Template,"AR by Insurance",ar_nc_df,len(template_ar_by_ins)+1,col_no1+1,col_no1+1)
Template = format_cells(Template,"AR by Insurance",len(template_ar_by_ins),len(template_ar_by_ins)+len(ar_nc_df),col_no+1,col_no1+1)
Template = bottom_border(Template,"AR by Insurance", len(template_ar_by_ins)+len(ar_nc_df),col_no+1,col_no1+1)
Template = left_border_one_cell(Template,"AR by Insurance", len(template_ar_by_ins)+len(ar_nc_df),col_no+1)
Template = right_border_one_cell(Template,"AR by Insurance", len(template_ar_by_ins)+len(ar_nc_df),col_no1+1)


#---------------------Yearly Fin------------------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet_df(template_month_fin_summary,"Facility",1)
to_row,to_col=coordinate_finder_insheet_df(template_month_fin_summary,"Net Change in Patient Balance",1)

#Paste values in excel
Template = copy_paste_raw(Template,"Monthly Fin Summary",year_fin_df,len(template_month_fin_summary)+1,col_no+1,to_col+1)
Template = format_cells(Template,"Monthly Fin Summary",len(template_month_fin_summary),len(template_month_fin_summary)+len(year_fin_df),col_no+1,to_col+1)
Template = bottom_border(Template,"Monthly Fin Summary", len(template_month_fin_summary)+len(year_fin_df),col_no+1,to_col+1)
Template = left_border_one_cell(Template,"Monthly Fin Summary", len(template_month_fin_summary)+len(year_fin_df),col_no+1)
Template = right_border_one_cell(Template,"Monthly Fin Summary", len(template_month_fin_summary)+len(year_fin_df),to_col+1)



#---------------------Collection Rate-------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet_df(template_collection_rate,"Facility",1)
to_row,to_col=coordinate_finder_insheet_df(template_collection_rate,"% of Adjustments",1)

#Paste values in excel
Template = copy_paste_raw(Template,"Collection Rate by Category",collection_df,len(template_collection_rate)+1,col_no+1,to_col+1)
Template = format_cells(Template,"Collection Rate by Category",len(template_collection_rate),len(template_collection_rate)+len(collection_df),col_no+1,to_col+1)
Template = bottom_border(Template,"Collection Rate by Category", len(template_collection_rate)+len(collection_df),col_no+1,to_col+1)
Template = left_border_one_cell(Template,"Collection Rate by Category", len(template_collection_rate)+len(collection_df),col_no+1)
Template = right_border_one_cell(Template,"Collection Rate by Category", len(template_collection_rate)+len(collection_df),to_col+1)


print("Pasting values and Formatting raw data tabs done")

Template.save(file_saving_location)
perform_calc(file_saving_location)

#Timer
end_time = time.time()
elapsed_time_seconds = end_time - start_time
elapsed_minutes = int(elapsed_time_seconds // 60)
elapsed_seconds = int(elapsed_time_seconds % 60)
print("Elapsed time: ", elapsed_minutes, "minutes", elapsed_seconds, "seconds")

