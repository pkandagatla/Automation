#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
from pandas.tseries.offsets import MonthEnd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import xlwings as xw
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.cell import get_column_letter
import getpass
import sys
import warnings
warnings.filterwarnings("ignore")


#Functions

#Reading raw data
def read_rawdata(filepath):
    app = xw.App(visible=False)
    book = xw.Book(filepath)
    if filepath[-3:] == "xls":
        filepath2 = filepath+"x"
        book.save(filepath2)
        os.remove(filepath)
        df = pd.read_excel(filepath2, index_col=None,header=None)
    else:
        df = pd.read_excel(filepath, index_col=None,header=None)
    book.close()
    app.kill()
    return(df)
    
def read_data(filepath):
    df = pd.read_excel(filepath, index_col=None,header=None)
    return(df)

#Get the Facility name
def facility_name(df,x,y):
    facility_name=df.iloc[x,y]
    return facility_name

#reordering_column
def reorder_column(df,columnname,position):
    df.insert(position,columnname,df.pop(columnname))
    return df

#To trim a column 
def trim_column(df,column_no):
    df.iloc[:,column_no] = df.iloc[:,column_no].str.strip()
    return df

def trim_many_columns(df,column_nos):
    for j in column_nos:
        for i in range(len(df)):
            if type(df.iloc[i,j]) != str:
                df.iloc[i,j] = str(df.iloc[i,j]).strip()
            else:
                df.iloc[i,j] = df.iloc[i,j].strip()
    return df

def trim_aging(df):
    for i in range(df.shape[0]):
        for j in range(2,df.shape[1]):
            df.iloc[i,j] = float(str(df.iloc[i,j]).replace(" ",'0'))
    return df

#get first row of dataframe
def first_rowofdf(df):
    first_row=pd.DataFrame()
    df=pd.DataFrame(df)
    first_row = df.iloc[0:1, :] 
    return first_row

#It slices the dataframe if it finds a value from "x" space of the dataframe
def iter_row_slicer(df,valuetofind,x):
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                df=df.iloc[i+x:].reset_index(drop=True)
                found=True
                break
        if found is True:
            break
    if(found is False):
        message="No Value"
        return message
    return df


#Here we find a particular values coordinate in a sheet with the recurrence at what time
def coordinate_finder_insheet(df,valuetofind,recurrence):   
    found=False
    count=0
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                row_no=i+1
                col_no=j+1
                count+=1
                if(count==recurrence):
                    found=True
                    break
        if found is True:
            break 
    return(row_no,col_no)
    
#Here we are pasting the values from a dataframe to the raw data sheet
def copy_paste_raw(Template,sheetname,df,from_row,from_col,to_col):
    
    Sheet=Template[sheetname]

    row_number=from_row
    i=0
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row-1+len(df),min_col=from_col,max_col=to_col):
        j=0
        col_number=from_col
        for cell in row:
            Sheet.cell(row=row_number,column=col_number).value=df.iloc[i,j]
            j+=1
            col_number+=1
        i+=1
        row_number+=1 
    return(Template)

#Here we are extending the formulas till the end we have the values
def extend_formulas(Template,sheetname,df,from_row,from_col,to_col):
    Sheet=Template[sheetname]
    i=from_row
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row-2+len(df),min_col=from_col,max_col=to_col):
        j=from_col
        for cell in row:
            Sheet.cell(row=i,column=j).value=Translator(Sheet.cell(row=from_row-1,column=j).value,origin=Sheet.cell(row=from_row-1,column=j).coordinate).translate_formula(Sheet.cell(row=i,column=j).coordinate)    
            j+=1
        i+=1
    return(Template)


def concatenate_dfs(dfs):
    df = pd.concat(dfs, ignore_index=True)
    return df
    
def first_rowofdf(df):
    first_row=pd.DataFrame()
    df=pd.DataFrame(df)
    first_row = df.iloc[0:1, :] 
    return first_row

def assign_value_to_df(df,x,y,value):
    df.iloc[x,y]=value
    return df

#function to delete values that exist after the particular value
def delete_after_slicer(df,valuetofind,x):   
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                df=df.iloc[0:i+x,:].reset_index(drop=True)  
                found=True
                break
        if found is True:
            break
    return df

#Triggers calculation in the sheet
def perform_calc(filepath):
    wb = xw.Book(filepath)
    app = xw.App(visible=False)
    wb.app.calculate()
    wb.save()
    wb.close()
    app.kill()
    

#Returns a dataframe converted from a excel sheet
def create_df_from_sheet(sheet_name,file_path):
    df=pd.read_excel(file_path,sheet_name=sheet_name)
    return df

#To get a unique value of the dataframe by giving the column indices
def unique_row_values(df,col_no):
    df=df.iloc[:,col_no].unique()
    df=pd.DataFrame(df)
    return df

def text_before(x,character):
    k = []
    for i in x :
        ind = i.find(character)
        k.append(i[0:ind])
    return k

def text_after(x,character):
    k = []
    for i in x :
        ind = i.find(character)
        k.append(i[ind:])
    return k

def unique_values_except(df, omit_list, column_index):
    wanted_list = []
    for i in range(len(df)):
        value = df.iloc[i, column_index]
        if value not in omit_list and value not in wanted_list:
            wanted_list.append(value)
    return wanted_list

#Deletes the below rows in the sheet 
def deleting_below_rows(worksheet,start_row,no_of_rows):
    worksheet.DeleteRow(start_row,no_of_rows)
    return worksheet
    
#Funtion to paste values
def paste_value_in_cell(workbook,sheet,row_no,col_no,paste_value):
    worksheet=workbook[sheet]
    worksheet.cell(row=row_no,column=col_no).value=paste_value
    return workbook

    
#To delete the specific cells in a range
def delete_cell_range(filepath,sheet_name,start_row,start_col,end_row,end_col):
    if(start_col<end_col ):
        workbook = Workbook()
        workbook.LoadFromFile(filepath)
        worksheet = workbook.Worksheets[sheet_name]   
        range_to_delete = worksheet.Range[start_row,start_col,end_row,end_col] 
        worksheet.DeleteRange(range_to_delete, DeleteOption.MoveLeft)
        workbook.SaveToFile(filepath)
        
def delete_columns(filepath,sheet_name,start_row,start_col,end_row,end_col):
    if(start_col<end_col ):
        workbook = Workbook()
        workbook.LoadFromFile(filepath)
        worksheet = workbook.Worksheets[sheet_name]   
        worksheet.DeleteColumn(start_col,end_col)
        workbook.SaveToFile(filepath)
        
#We are iterating to the row to find how many times we have a particular value and if we need it 8 times the 8th found value coordinates is stored
def how_many_values_finder_insheet(template_path,sheetname,valuetofind,no_payers):
    times_recurring=0
    found=False
    df=pd.read_excel(template_path,sheet_name=sheetname,header=None)
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if((df.iloc[i,j]==valuetofind) and (times_recurring<=no_payers)):
                row_no=i+1
                col_no=j+1
                times_recurring+=1
            if(times_recurring==no_payers):
                found=True
                break
        if  found is True:
            break          
    return(row_no,col_no)

#Here we are finding the coordinate with a bound in excel sheet
def coordinate_finder_special(Template,sheetname,value_to_find,from_row,from_col,to_row,to_col):
    
    Sheet=Template[sheetname]
    found=False
    row_number=from_row
    i=0
    for row in Sheet.iter_rows(min_row=from_row,max_row=to_row,min_col=from_col,max_col=to_col):
        j=0
        col_number=from_col
        for cell in row:
            if(Sheet.cell(row=row_number,column=col_number).value==value_to_find):
                found=True
                break
            j+=1
            col_number+=1
        if found is True:
            break
        i+=1
        row_number+=1 
    Template.close()
    return (row_number,col_number)

def to_find_no_months(start_date,end_date):
    dif=relativedelta(end_date,start_date)
    no_months= dif.years*12 + dif.months+1
    return no_months

#Deletes the below rows in the sheet 
def deleting_right_columns(worksheet,start_col,no_of_cols):
    worksheet.DeleteColumn(start_col,no_of_cols)
    return worksheet

#Will add a bottom border for a row in a column range
def bottom_border_one_cell(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def right_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_right_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def right_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def top_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style='medium'), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_right_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template
def get_cell_range(start_row, start_col, end_row, end_col):
    # Convert column indices to letters
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(end_col)
    
    # Construct the cell range string
    cell_range = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
    return cell_range

def cut_cell_range(Template,sheet_name,cell_range,row_to_move,col_to_move):
    ws=Template[sheet_name]
    ws.move_range(cell_range, rows=row_to_move, cols=col_to_move)
    return Template

def cut_cell_range_translate(Template,sheet_name,cell_range,row_to_move,col_to_move):
    ws=Template[sheet_name]
    ws.move_range(cell_range, rows=row_to_move, cols=col_to_move,translate=True)
    return Template

#Here we find a particular values coordinate in a sheet with the recurrence at what time
def coordinate_finder_insheet_df(df, valuetofind, recurrence):   
    found = False
    count = 0
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if df.iloc[i, j] == valuetofind:
                row_no = i
                col_no = j
                count += 1
                if count == recurrence:
                    found = True
                    break
        if found:
            break
    if found:
        return (row_no, col_no)
    else:
        return (0, 0) 
    

import time
start_time = time.time()

# Get the current system's username
username = getpass.getuser()
print("Username:", username)

payer_grp_mail = f"C:/Users/{username}/Downloads/AR Assessment - Payer Grouping.xlsx"

#New user need to store this
path_folder = r"C:\Users\pragna_kandagatla\Desktop\Automation_LPC_Multifacility"

#Checking the files required i.e setup files required and storing
for filename in os.listdir(path_folder):
    file_path4 = os.path.join(path_folder, filename)
    if filename.endswith("AR Assessment - Multi Facility Template (LPC).xlsx"):
        template_path = file_path4
    elif filename.endswith("AR Assessment - Payer Grouping.xlsx"):
        payer_grp_send = file_path4

folder_path = sys.argv[1]
folder_path=folder_path.replace("\\","/").replace("\"","")

if __name__ == "__main__":
    folder_path=input("Paste the path of raw data folder= ")
    folder_path=folder_path.replace("\\","/").replace("\"","")

folder_name=folder_path.split("/")[-1]
file_saving_location = f"C:/Users/{username}/Downloads/AR Assessment - {folder_name}.xlsx"

files_read_check = 1
for j, filename2 in enumerate(os.listdir(folder_path)):
    file_path2 = os.path.join(folder_path, filename2)
    if os.path.isfile(file_path2):
        file_path2=file_path2.replace("\\","/").replace("\"","")
        read_df = pd.read_excel(file_path2, header=None, engine="openpyxl")
        row_no_payer, col_payer = coordinate_finder_insheet_df(read_df, "Payer", 1)
        if read_df.iloc[row_no_payer, col_payer + 1] == "Payer Group":
            payer_group = file_path2
            files_read_check += 1
            payer_grp_check="Yes"
    else:
            
        for i, filename in enumerate(os.listdir(file_path2)):
            if i > 0: break
            file_path = os.path.join(file_path2, filename)
            file_path=file_path.replace("\\","/").replace("\"","")
            read_df = read_rawdata(file_path)
            read_df = trim_column(read_df,0)
            row_no_aging, col_aging = coordinate_finder_insheet_df(read_df, "Resident Name", 1)
            row_no_ar, col_ar = coordinate_finder_insheet_df(read_df, "Billing Period", 1)
            row_no_cash, col_cash = coordinate_finder_insheet_df(read_df, "GL Account Description", 1)

            locals()[f"excel{j}"]  =read_df
            if read_df.iloc[row_no_aging + 1, col_aging] == "Payor Type":
                Aging_path = file_path2
                files_read_check += 1
            elif read_df.iloc[row_no_ar, col_ar+1] == "GL Posting Period":
                AR_path = file_path2
                files_read_check += 1 
            elif read_df.iloc[row_no_cash + 1, col_cash] == "Resident Name":
                Cash_path = file_path2
                files_read_check += 1

if not(files_read_check>4):
    print(f"Some of the raw data files are missing in this folder = {folder_path}")
    sys.exit()
# After the loop, you can check the files assigned
print("Files assigned:")
print(f"Aging Report: {Aging_path}")
print(f"AR Report: {AR_path}")
print(f"Cash Report: {Cash_path}")
print(f"Payer group: {payer_group}")

#Loading workbook
Template=load_workbook(filename=template_path)

#storing all the sheets as a dataframe
template_aging=pd.read_excel(template_path,sheet_name="Aging",header=None)
template_ar=pd.read_excel(template_path,sheet_name="AR Detail",header=None)
template_cash=pd.read_excel(template_path,sheet_name="Cash",header=None)
template_grouping=pd.read_excel(template_path,sheet_name="Payer Grouping",header=None)
template_key=pd.read_excel(template_path,sheet_name="Key",header=None)
template_ncr_cal=pd.read_excel(template_path,sheet_name="Payer-Wise NCR Cal",header=None)
template_aging_cal=pd.read_excel(template_path,sheet_name="Aging Cal",header=None)
template_clean_up=pd.read_excel(template_path,sheet_name="Clean Up estimates",header=None)
template_ncr=pd.read_excel(template_path,sheet_name="NCR",header=None)
template_assessment_metrics=pd.read_excel(template_path,sheet_name="Assessment Metrics",header=None)
template_ncr_facility=pd.read_excel(template_path,sheet_name="NCR by Facility",header=None)
template_ncr_payer=pd.read_excel(template_path,sheet_name="NCR by Payer",header=None)

print("Template files are converted to dataframe")

#-----------------Aging----------

Aging_df = pd.DataFrame()
aging_facility =[]
Month_list = []
Aging_payers_list = {}
Aging_date_range = {}

for filename in os.listdir(Aging_path):
    file_path = os.path.join(Aging_path, filename)
    
    #Passing Aging report file path to retrieve dataframe of it
    Aging_report_df=read_rawdata(file_path)

    #trim first column
    Aging_report_df= trim_column(Aging_report_df,0)
    Aging_report_df2=Aging_report_df

    #Get facility name
    A_Facility_Name=facility_name(Aging_report_df,0,0)
    Aging_report_df["Facility Name"]=A_Facility_Name
    aging_facility.append([filename.split('-',1)[0],A_Facility_Name])
    
    #reorder column
    Aging_report_df=reorder_column(Aging_report_df,"Facility Name",0)

    #Slice for Resident aging here we need first column to know the aging buckets
    Aging_report_df=iter_row_slicer(Aging_report_df,"Resident Name",0)
    Aging_report_df_age_bucket =iter_row_slicer(Aging_report_df,"Payor Type",0)

    # First row
    Month=pd.DataFrame()
    Month=first_rowofdf(Aging_report_df)
    Month_list.append(Month.iloc[0])
    Age_bucket=pd.DataFrame()
    Age_bucket=first_rowofdf(Aging_report_df_age_bucket)

    #Initial Raw Data Checks - checking number of columns
    check=0
    Aging_report_dfheaders=["Facility_Name","Payor Type","Total Due","Current","1 Month","2 Months","3 Months","4 Months","5 Months",                            "6 Months","7 Months","8 Months","9 Months","10 Months","11 Months","12 Months","12+ Months"]

    for i in range(1,len(Aging_report_dfheaders)):
        if(Aging_report_dfheaders[i]==Age_bucket.iloc[0,i].strip()):
            check+=1
    if(check!=16):
        print(Age_bucket.iloc[0])   
        print("Aging buckets are in unusual format")

    # Checking for Payor Type Summary
    aging_check=iter_row_slicer(Aging_report_df2,"Payor Type Summary",1)
    if (type(aging_check) == str) | (aging_check.shape[0]<4) | (aging_check.iloc[-1,1] != "Report Totals") :
        print("We don't have Payor Type Summary for the Assessment")
        
    Aging_report_df=iter_row_slicer(Aging_report_df,"Payor Type Summary",2)
    Aging_report_df = Aging_report_df.fillna(0)
    Aging_report_df = trim_aging(Aging_report_df)

    #function to delete values that exist after the particular value
    Aging_report_df=delete_after_slicer(Aging_report_df,"Report Totals",1)
    temp = Aging_report_df.iloc[2:,2:].replace(" ","")
    temp = temp.replace(np.nan,0)
    Aging_report_df.iloc[2:,2:] = temp
        
    Aging_report_df["filename"] = filename.split('-',1)[0]
        
    #Concatenate all dataframe
    Aging_df=concatenate_dfs([Aging_df,Aging_report_df])
        
    aging_end_index = pd.to_datetime(Aging_report_df2.iloc[1,4].strip().split(" ")[-1])
    Aging_date_range[filename.split('-',1)[0]] = aging_end_index
        
    #Aging Payers
    mask = Aging_report_df[0] == 0
    Aging_report_df = Aging_report_df[~mask]
    Aging_payers = Aging_report_df[0]
    Aging_payers_df = Aging_payers.dropna()
    Aging_payers_df = Aging_payers_df.unique()
    Aging_payers_df  = list(filter(lambda x: x.find('Resident Name') == -1 , Aging_payers_df))
    Aging_payers_df  = list(filter(lambda x: x.find('Payor Type') == -1 , Aging_payers_df))
    Aging_payers_df  = list(filter(lambda x: x.find('Report Totals') == -1 , Aging_payers_df))
    Aging_payers_df=list(filter(lambda x: len(x)>0, Aging_payers_df))
    Aging_payers_list[filename.split('-',1)[0]] = Aging_payers_df

Header_df=concatenate_dfs([Month,Age_bucket])
Aging_df=concatenate_dfs([Header_df, Aging_df])
    
#Give values to some postion in dfs
Aging_df=assign_value_to_df(Aging_df,0,0,"")
Aging_df=assign_value_to_df(Aging_df,1,0,"Facility Name")
    
aging_facility = pd.DataFrame(aging_facility,columns=["file","facility"])


#-------------------------------Cash-------------------------------------------

Cash_df = pd.DataFrame()
cash_facility =[]
Cash_date_range = {}

# Predefined column names
CreportCnames = ["Facility_Name","Resident Name","ID" ,"Trans Date","To Mnth", "Count", "Unit Amount", "Extended Amount", "Bill Code", "Empty1","Private Pay Type","Private Amount","3rd Party Type","3rd Party Amount","Empty2"]
Creportheaders = ["Resident Name","ID" ,"Trans Date","To Mnth", "Count", "Amount", "Amount", "Code", "","Type","Amount","Type","Amount",""]

for filename in os.listdir(Cash_path):
    file_path = os.path.join(Cash_path, filename)
    #Passing Cash report file path to retrieve dataframe of it 
    Cash_report_df=read_rawdata(file_path)

    #trim first column
    Cash_report_df= trim_column(Cash_report_df,0)
    Cash_report_df2=Cash_report_df
    Cash_report_df2 = Cash_report_df2.fillna("")

    #Initial Raw Data checks - checking number of columns
    check = 0
    for i in range(len(Creportheaders)):
        if(Creportheaders[i]==Cash_report_df2.iloc[6,i].strip()):
            check+=1
    if(check!=14):
        print(Cash_report_df2.iloc[6])
        print("Column names does not match for cash report")
        
    C_Facility_Name=facility_name(Cash_report_df,0,0)
    Cash_report_df["Facility Name"]=C_Facility_Name
    cash_facility.append([filename.split('-',1)[0],C_Facility_Name])

    #reorder column
    Cash_report_df=reorder_column(Cash_report_df,"Facility Name",0)

    #Headers
    Cash_report_df = Cash_report_df.set_axis(CreportCnames,axis=1)

    #Row slice if it finds the Deposit distribution
    Cash_report_df=iter_row_slicer(Cash_report_df,"GL Account Description",4)

    #function to delete values that exist after the particular value
    Cash_report_df=delete_after_slicer(Cash_report_df,"Payor Type Summary",1)

    #trim columns
    Cash_report_df = trim_many_columns(Cash_report_df,[8,10,12])
    Cash_report_df = Cash_report_df.replace("nan","")
    #Concatenate all dataframe
    Cash_df=concatenate_dfs([Cash_df,Cash_report_df])
        
    c_start_index = pd.to_datetime(Cash_report_df2.iloc[2,0].split(" ")[7]) + MonthEnd(0)
    c_end_index = pd.to_datetime(Cash_report_df2.iloc[2,0].split(" ")[9]) + MonthEnd(0)
    c_date_range = str(c_start_index)[:10] + " - " + str(c_end_index)[:10]
    Cash_date_range[filename.split('-',1)[0]] = c_date_range

cash_facility = pd.DataFrame(cash_facility,columns=["file","facility"])  


#-------------------------------------AR Detail-----------------------------------

AR_df = pd.DataFrame()
k = 0
AR_date_range = {}

# Predefined column names
ARreportCnames = ["Billing Period","GL Posting Period" ,"Profile ID","Transaction Source", "Billing Period-Apply", "Transaction Date",                 "Facility Type", "Census Payor Type", "Billing Code","Description","Contractual Allow","Count","Unit Amount",                "Bill Amount","Private Pay Amount", "Private Payor Type","Third Party Amount","Third Party Payor Type",                  "Third Party Bill Type","From Date", "Thru Date"]

for filename in os.listdir(AR_path):
    file_path = os.path.join(AR_path, filename)

    #Passing AR report file path to retrieve dataframe of it 
    AR_report_df=read_rawdata(file_path)

    #Initial Raw Data checks - checking number of columns
    check = 0
    for i in range(len(ARreportCnames)):
        if(ARreportCnames[i]==AR_report_df.iloc[0,i].strip()):
            check+=1
    if(check!=21):
        print(AR_report_df.iloc[0])
        print("Column names does not match for AR report")
        
    #Headers
    AR_report_df = AR_report_df.set_axis(ARreportCnames,axis=1)

    #Get facility name
    ar_facility = filename.split('-',1)[0]
    AR_report_df["Facility Name"] = cash_facility[cash_facility["file"] == ar_facility]['facility'][k]
    k+=1
        
    #reorder column
    AR_report_df=reorder_column(AR_report_df,"Facility Name",0)
    AR_report_df = trim_many_columns(AR_report_df,[9,10,16,18])
    AR_report_df = AR_report_df.replace("nan","")
    AR_report_df = AR_report_df.iloc[1:,:]
        
    #Concatenate all dataframe
    AR_df=concatenate_dfs([AR_df,AR_report_df])
        
    #Date
    a_start_index=AR_report_df.iloc[0,1]
    a_end_index=AR_report_df.iloc[-1,1]
    a_date_range = str(a_start_index)[:10] + " - " + str(a_end_index)[:10]
    AR_date_range[filename.split('-',1)[0]] = a_date_range


#------------------------------------RAW DATA CHECKS---------------------------------

aging_facility = aging_facility.sort_values(by="file")
cash_facility = cash_facility.sort_values(by="file")

#Facility Name check
check = 0
for i in range(len(aging_facility)):
    if aging_facility["facility"][i] == cash_facility["facility"][i]:
        check+=1 
        
if check != len(aging_facility):
    print("Facility Names are not matching between reports")
    
    for i in range(len(aging_facility)):
        print("Aging Report= ",aging_facility["facility"][i],"\nCash Report= ",cash_facility["facility"][i],"\n")
        Aging_df.loc[(Aging_df["filename"] == cash_facility["file"][i]),"Facility Name"] = cash_facility["facility"][i]
    Aging_df = Aging_df.drop(["filename"],axis=1)
               
#Date Range Check
if len(set(Aging_date_range.values())) == len(set(Cash_date_range.values())) == len(set(AR_date_range.values())) == 1:
    if set(Cash_date_range.values()) == set(AR_date_range.values()):
        if list(set(AR_date_range.values()))[0].split(" - ")[1] == str(list(set(Aging_date_range.values()))[0]).split(" ")[0]:
            print("Date Ranges are matching between reports")
else:
    print("We have different Date Ranges for the reports \n Cash Report= ",list(set(Cash_date_range.values()))[0],          "\n AR Detail Report= ", list(set(AR_date_range.values()))[0],"\n Aging Report= ",          str(list(set(Aging_date_range.values()))[0]).split(" ")[0])
        
print("Performed all raw data checks")



#--------------------------------------Payers Extraction---------------------------------------------------
mask = Aging_df[0] == 0
Aging_df = Aging_df[~mask]
Aging_payers = Aging_df[0]
Aging_payers_df = Aging_payers.dropna()

Cash_payers_priv = Cash_df['Private Pay Type']
Cash_payers_third = Cash_df['3rd Party Type']
AR_payers_priv = AR_df['Private Payor Type']
AR_payers_third = AR_df['Third Party Payor Type']

payers_df = concatenate_dfs([Cash_payers_priv,Cash_payers_third,AR_payers_priv,AR_payers_third])
payers_df = payers_df.dropna()

unique_payers_ncr = payers_df.unique()
unique_payers_ncr  = list(filter(lambda x: x.find('MAGIC') == -1 , unique_payers_ncr))
unique_payers_ncr=list(filter(lambda x: len(x)>0, unique_payers_ncr))

Aging_payers_df = Aging_payers_df.unique()
Aging_payers_df  = list(filter(lambda x: x.find('Resident Name') == -1 , Aging_payers_df))
Aging_payers_df  = list(filter(lambda x: x.find('Payor Type') == -1 , Aging_payers_df))
Aging_payers_df  = list(filter(lambda x: x.find('Report Totals') == -1 , Aging_payers_df))
Aging_payers_df=list(filter(lambda x: len(x)>0, Aging_payers_df))

Aging_payers_payortype = text_before(Aging_payers_df," ")
Aging_payers_payor = Aging_payers_df
Aging_payers_final = pd.DataFrame({'Payor Type' : Aging_payers_payortype,
                                'Payor' : Aging_payers_payor })

Aging_payers_final = trim_column(Aging_payers_final,1)
ncr_payers_df = pd.DataFrame({'Payor Type' : unique_payers_ncr})

payers_list = pd.merge(Aging_payers_final, ncr_payers_df, on='Payor Type', how='outer')

#Pasting the raw data in the Payer group template
payer_grp_file=load_workbook(payer_grp_send)

#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_grouping,"Payer",1)

#Paste values in excel
payer_grp_file=copy_paste_raw(payer_grp_file,"Payer Grouping",payers_list,row_no+2,col_no+1,col_no+2)

#Saving the file
payer_grp_file.save(payer_grp_mail)
perform_calc(payer_grp_mail)

print("Payer Group file is Generated!")
   

#--------------------------------Payer Group file upload------------------------------------------
if(payer_grp_check=="Yes"):
        
    #Payer Group Dataframe
    payer_group_df= read_data(payer_group)
    payer_group_df.columns=payer_group_df.iloc[1]

    #Slice for Payer
    payer_group_df=iter_row_slicer(payer_group_df,"Payer",1)
    payer_group_df=payer_group_df.iloc[:,1:4]
    
    
    payer_group_df_paste=payer_group_df.iloc[:,:3]

    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_grouping,"Payer",1)

    #Paste values in excel
    Template=copy_paste_raw(Template,"Payer Grouping",payer_group_df_paste,row_no+1,col_no,col_no+2)
    
    #Payer Grouping copy paste
    row_no, col_no=coordinate_finder_insheet(template_grouping,"Payer Groups",1)
    
    #Getting value for number of payers in the assessment
    payer_group_df.loc[-1] = ["", "", "Private"]  
    payer_group_df.index = payer_group_df.index + 1 
    payer_group_df = payer_group_df.sort_index()
    payer_group_df=trim_column(payer_group_df,2)
    unique_payer_groups=unique_values_except(payer_group_df,["Non-QHCR Billed Payer"],2)
    unique_payer_groups = pd.DataFrame(unique_payer_groups)
    unique_facility_groups=unique_row_values(cash_facility,1)

    #Paste values in excel
    Template=copy_paste_raw(Template,"Payer Grouping",unique_payer_groups,row_no+1,col_no,col_no)
    
    #Facility copy paste
    row_no, col_no=coordinate_finder_insheet(template_grouping,"Facility List",1)
    Template=copy_paste_raw(Template,"Payer Grouping",unique_facility_groups,row_no+1,col_no,col_no)
    
    #Clean-up estimates copy paste
    #Unique list of payers need to be added in the AR Clean up estimates
    wanted_list = unique_values_except(payer_group_df, ["Private", "Other", "Miscellaneous","Non-QHCR Billed Payer"], 2)
    wanted_list=pd.DataFrame(wanted_list)
    row_no, col_no=coordinate_finder_insheet(template_grouping,"Payer Group - AR Clean up",1)
    Template=copy_paste_raw(Template,"Payer Grouping",wanted_list,row_no+1,col_no,col_no)
    
print("Payer Grouping tab updated")


#------------------------------------------Calculation Sheets--------------------------------------

no_facility = len(aging_facility)
no_of_clean_up_payers = len(wanted_list)
no_of_payer_grps = len(unique_payer_groups)
no_months = to_find_no_months(AR_df["Billing Period"].min(),AR_df["Billing Period"].max())


#----------------------------------------------NCR-----------------------------------
#to get number of months

#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_ncr,"Metrics",1)

#To paste the end range month for assessment period
Template=paste_value_in_cell(Template,"NCR",row_no,col_no+19,a_end_index)

#-----------------------------------------Aging Cal------------------------------------------

for i in range(len(aging_facility)):
    #Getting position
    row_num, col_num=coordinate_finder_insheet(template_aging_cal,"Payor Type",i+1)
    row_end, col_end=coordinate_finder_insheet(template_aging_cal,"Payer Total",i+1)
    
    payers = pd.DataFrame(Aging_payers_list[aging_facility["file"][i]])
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Aging Cal",payers,row_num+1,col_num,col_num)

Template.save(file_saving_location)
print("Values needed to be pasted in Calculation sheets are done!")
print("Deletion of rows/cells started......")


#--------------------------------Deletion of Extra rows for Calculation sheets Start---------------------
#creating variables for Spire libraries worksheets
from spire.xls.common import *
from spire.xls import *

workbook = Workbook()
workbook.LoadFromFile(file_saving_location)

worksheet_aging_cal = workbook.Worksheets["Aging Cal"]   
worksheet_clean_up= workbook.Worksheets["Clean Up estimates"] 
worksheet_ncr_cal= workbook.Worksheets["Payer-Wise NCR Cal"] 
worksheet_ncr_facility= workbook.Worksheets["NCR By Facility"] 
worksheet_ncr_payer= workbook.Worksheets["NCR By Payer"] 
worksheet_ar_assessment= workbook.Worksheets["Assessment Metrics"]
worksheet_ncr= workbook.Worksheets["NCR"]

# ----------------------Deletion for Aging Cal-------------------
#Find the position of the particular excel paths iterating over each table and deleting the extra payers here
row_reducing=0

row_no,col_no=coordinate_finder_insheet(template_aging_cal,"Payor Group Summary",1)
end_row_no,end_col_no=coordinate_finder_insheet(template_aging_cal,"Payer Type Total",1)
    
start_row=row_no+no_of_payer_grps+1-row_reducing
no_of_rows_delete=end_row_no-start_row-row_reducing
    
worksheet_aging_cal=deleting_below_rows(worksheet_aging_cal,start_row,no_of_rows_delete)
row_reducing+=no_of_rows_delete

for i in range(1,no_facility+1):
    row_no,col_no=coordinate_finder_insheet(template_aging_cal,"Payor Type",i)
    end_row_no,end_col_no=coordinate_finder_insheet(template_aging_cal,"Payer Total",i)
    
    start_row=row_no+len(Aging_payers_list[aging_facility["file"][i-1]])+1-row_reducing
    no_of_rows_delete=end_row_no-start_row-row_reducing
    
    worksheet_aging_cal=deleting_below_rows(worksheet_aging_cal,start_row,no_of_rows_delete)
    row_reducing+=no_of_rows_delete
    
#Deleting the Aging Cal extra facilities
row_no_fac,col_no_fac=coordinate_finder_insheet(template_aging_cal,"Payor Type",no_facility+1)

start_row_fac=row_no_fac-row_reducing
worksheet_aging_cal=deleting_below_rows(worksheet_aging_cal,start_row_fac,10000)
print("Aging Cal deletion completed")


# ----------------------Deletion for Clean up estimates-------------------
#Find the position of the particular excel paths iterating over each table and deleting the extra payers here
row_reducing=0

row_no,col_no=coordinate_finder_insheet(template_clean_up,"Payer Type",1)
end_row_no,end_col_no=coordinate_finder_insheet(template_clean_up,"Payer Type Total",1)
    
start_row=row_no+no_of_clean_up_payers+1-row_reducing
no_of_rows_delete=end_row_no-start_row-row_reducing

worksheet_clean_up=deleting_below_rows(worksheet_clean_up,start_row,no_of_rows_delete)
row_reducing+=no_of_rows_delete 

row_no,col_no=coordinate_finder_insheet(template_clean_up,"Facility wise Summary",i)
end_row_no,end_col_no=coordinate_finder_insheet(template_clean_up,"Facility Wise Totals",i)
    
start_row=row_no+no_facility+1-row_reducing
no_of_rows_delete=end_row_no-start_row-row_reducing

worksheet_clean_up=deleting_below_rows(worksheet_clean_up,start_row,no_of_rows_delete)
row_reducing+=no_of_rows_delete

for i in range(2,no_facility+1+1):
    row_no,col_no=coordinate_finder_insheet(template_clean_up,"Payer Type",i)
    end_row_no,end_col_no=coordinate_finder_insheet(template_clean_up,"Payer Type Total",i)
    
    start_row=row_no+no_of_clean_up_payers+1-row_reducing
    no_of_rows_delete=end_row_no-start_row-row_reducing
    
    worksheet_clean_up=deleting_below_rows(worksheet_clean_up,start_row,no_of_rows_delete)
    row_reducing+=no_of_rows_delete
    
    
#Deleting the Clean up extra facilities
row_no_fac,col_no_fac=coordinate_finder_insheet(template_clean_up,"Payer Type",no_facility+2)

start_row_fac=row_no_fac-row_reducing
worksheet_clean_up=deleting_below_rows(worksheet_clean_up,start_row_fac,10000)
print("AR Clean up estimates deletion completed")

# ----------------------Deletion for NCR Payer-------------------

#Deletion for the number of months valid in the assessment
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_ncr_payer,"Facility",1)

# worksheet_ncr_payer=delete_cell_range(worksheet_ncr_payer,row_no,col_no+1,10000,col_no+18-no_months)
no_of_cols_to_delete=18-no_months
if no_of_cols_to_delete>=0:
    worksheet_ncr_payer=deleting_right_columns(worksheet_ncr_payer,col_no+1,no_of_cols_to_delete)

#Find the position of the particular excel paths iterating over each table and deleting the extra payers here
row_reducing=0

#first file is the number of facilities variable

for i in range(1,no_of_payer_grps+1):
    row_no,col_no=coordinate_finder_insheet(template_ncr_payer,"Facility",i)
    end_row_no,end_col_no=coordinate_finder_insheet(template_ncr_payer,"Overall",i)
    
    start_row=row_no+no_facility+1-row_reducing
    no_of_rows_delete=end_row_no-start_row-row_reducing
    
    worksheet_ncr_payer=deleting_below_rows(worksheet_ncr_payer,start_row,no_of_rows_delete)
    row_reducing+=no_of_rows_delete
    
#Deleting the NCR Payer extra facilities
row_no_fac,col_no_fac=coordinate_finder_insheet(template_ncr_payer,"Facility",no_of_payer_grps+1)

start_row_fac=row_no_fac-row_reducing
worksheet_ncr_payer=deleting_below_rows(worksheet_ncr_payer,start_row_fac,10000)

print("NCR Payer wise deletion completed")

#----------------NCR Facility Wise--------------------------
#Deletion for the number of months valid in the assessment
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_ncr_facility,"Payer Type",1)

no_of_cols_to_delete=18-no_months
if no_of_cols_to_delete>=0:
    worksheet_ncr_facility=deleting_right_columns(worksheet_ncr_facility,col_no+1,no_of_cols_to_delete)

#Find the position of the particular excel paths iterating over each table and deleting the extra payers here
row_reducing=0

#no_facility is the number of facilities variable

for i in range(1,no_facility+1):
    row_no,col_no=coordinate_finder_insheet(template_ncr_facility,"Payer Type",i)
    end_row_no,end_col_no=coordinate_finder_insheet(template_ncr_facility,"Overall",i)
    
    start_row=row_no+no_of_payer_grps-row_reducing+1
    no_of_rows_delete=end_row_no-start_row-row_reducing
    
    worksheet_ncr_facility=deleting_below_rows(worksheet_ncr_facility,start_row,no_of_rows_delete)
    row_reducing+=no_of_rows_delete
    
#Deleting the NCR facility extra facilities
row_no_fac,col_no_fac=coordinate_finder_insheet(template_ncr_facility,"Payer Type",no_facility+1)

start_row_fac=row_no_fac-row_reducing
worksheet_ncr_facility=deleting_below_rows(worksheet_ncr_facility,start_row_fac,10000)
print("NCR Facility wise deletion completed")

# ----------------------Deletion for NCR Cal-------------------
#Deletion for the number of months valid in the assessment
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_ncr_cal,"Metrics",1)

no_of_cols_to_delete=18-no_months
if no_of_cols_to_delete>=0:
    worksheet_ncr_cal=deleting_right_columns(worksheet_ncr_cal,col_no+1,no_of_cols_to_delete)


#Find the position of the particular excel paths iterating over each table and deleting the extra payers here
row_reducing=0

row_no_initial,col_no_initial=coordinate_finder_insheet(template_ncr_cal,"Metrics",1)
row_no_payer_initial,col_no_payer_initial=coordinate_finder_insheet(template_ncr_cal,"Expected Revenue - Overall",no_of_payer_grps+1)
rows_to_keep=row_no_payer_initial-row_no_initial

for i in range(1,no_facility+1+1):
    row_no,col_no=coordinate_finder_insheet(template_ncr_cal,"Metrics",i)
    end_row,end_col=coordinate_finder_insheet(template_ncr_cal,"Metrics",i+1)
       
    start_row=row_no+rows_to_keep-row_reducing
    no_of_rows_delete=end_row-start_row-row_reducing-1

    worksheet_ncr_cal=deleting_below_rows(worksheet_ncr_cal,start_row,no_of_rows_delete)
    row_reducing+=no_of_rows_delete

# #Deleting the extra facilities
row_no_fac,col_no_fac=coordinate_finder_insheet(template_ncr_cal,"Metrics",no_facility+1+1)
start_row_fac=row_no_fac-row_reducing
worksheet_ncr_cal=deleting_below_rows(worksheet_ncr_cal,start_row_fac,10000)

print("Deletion of NCR Cal completed")

#--------AR Assessment Metrics------------
row_no,col_no=coordinate_finder_insheet(template_assessment_metrics,"Parameters",1)

start_col=col_no+2+no_facility
worksheet_ar_assessment=deleting_right_columns(worksheet_ar_assessment,start_col,30)

#--------NCR------------

#Deletion for the number of months valid in the assessment
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_ncr,"Metrics",1)

no_of_cols_to_delete=18-no_months
if no_of_cols_to_delete>=0:
    worksheet_dsoncr=deleting_right_columns(worksheet_ncr,col_no+1,no_of_cols_to_delete)

#Deletion of extra facilities
row_no,col_no=coordinate_finder_insheet(template_ncr,"Metrics",no_facility+2)

start_row=row_no

worksheet_ncr=deleting_below_rows(worksheet_ncr,start_row,100000)

print("Deletion of other output tabs completed")

workbook.SaveToFile(file_saving_location)
workbook.Dispose()
print("Triggering Calculation")
perform_calc(file_saving_location)

#------------------------------------Pasting Raw Data in Excel-----------------------------------

Template = load_workbook(filename=file_saving_location)

#----------------------CashReport-------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_cash,"Facility",1)
to_row,to_col=coordinate_finder_insheet(template_cash,"Empty#",1)

#Paste values in excel
Template =copy_paste_raw(Template,"Cash",Cash_df,row_no+2,col_no,to_col)

#Formula pasting range finder
for_row_no,for_col_no=coordinate_finder_insheet(template_cash,"EO month of Apply to Mnth date",1)
for_to_row,for_to_col=coordinate_finder_insheet(template_cash,"Payer Group - Private Type",1)

#Extend Formulas in Sheets
Template=extend_formulas(Template,"Cash",Cash_df,for_row_no+2,for_col_no,for_to_col)


#------------------------------AR Detail Report---------------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_ar,"Facility",1)
to_row,to_col=coordinate_finder_insheet(template_ar,"Thru Date",1)

#Paste values in excel
Template =copy_paste_raw(Template,"AR Detail",AR_df,row_no+1,col_no,to_col)

#Formula pasting range finder
for_row_no,for_col_no=coordinate_finder_insheet(template_ar,"EO Month Billing period-Apply date",1)
for_to_row,for_to_col=coordinate_finder_insheet(template_ar,"Payer Group - Private Type",1)

#Extend Formulas in Sheets
Template=extend_formulas(Template,"AR Detail",AR_df,for_row_no+2,for_col_no,for_to_col)


#----------------------------------------Aging----------------------------------------

#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_aging,"Facility",1)
to_row,to_col=coordinate_finder_insheet(template_aging,"12+ Months",1)

#Paste values in excel
Template=copy_paste_raw(Template,"Aging",Aging_df,row_no-1,col_no,to_col)


#------------------------------------------Cash Key---------------------------------------
Cash_key = Cash_df['Bill Code'].unique()
Cash_key_df = pd.DataFrame({"Bill Code":Cash_key})
Cash_key_df['Bill Code'] = Cash_key_df['Bill Code'].replace('', np.nan)
Cash_key_df = Cash_key_df.dropna()
Cash_key_df['Include'] = [1 if x == 'PAY' else 0 for x in Cash_key_df['Bill Code']]

#Cash key copy paste
row_no, col_no=coordinate_finder_insheet(template_key,"Bill Code",1)

Template=copy_paste_raw(Template,"Key",Cash_key_df,row_no+1,col_no,col_no+1)

#-------------------------------------AR key---------------------------------------------------
AR_key_df = AR_df.loc[:,['Billing Code','Description']]
AR_key_df = AR_key_df.drop_duplicates()

AR_key_df["Include"] = 1
AR_key_df["Bad debt"] = 0
AR_key_df = AR_key_df.reset_index()

for i in range(len(AR_key_df)):
    if AR_key_df["Billing Code"][i] == "PAY":
        AR_key_df["Include"][i] = 0
    elif (AR_key_df["Description"][i].lower()).find("refund") != -1:
        AR_key_df["Include"][i] = 2
    elif (AR_key_df["Description"][i].lower()).find("bad debt") != -1:
        AR_key_df["Include"][i] = 3
        AR_key_df["Bad debt"][i] = 1
    elif ((AR_key_df["Description"][i].lower()).find("c/a") != -1) or ((AR_key_df["Description"][i].lower()).find("c/adj") != -1) or ((AR_key_df["Description"][i].lower()).find("contractual adjustment") != -1):
        AR_key_df["Include"][i] = 3
        AR_key_df["Bad debt"][i] = 2
    elif ((AR_key_df["Description"][i].lower()).find("w/o") != -1) or ((AR_key_df["Description"][i].lower()).find("write-off") != -1) or ((AR_key_df["Description"][i].lower()).find("write off") != -1):
        AR_key_df["Include"][i] = 3
        
AR_key_df = AR_key_df.drop("index",axis=1)

#AR key copy paste
row_no, col_no=coordinate_finder_insheet(template_key,"Billing Code",1)

Template=copy_paste_raw(Template,"Key",AR_key_df,row_no+1,col_no,col_no+3)

print("Key tab updated")
print("Formatting of cells started......")


border_ar=pd.read_excel(file_saving_location,sheet_name="AR Detail",header=None)
border_cash=pd.read_excel(file_saving_location,sheet_name="Cash",header=None)
border_aging=pd.read_excel(file_saving_location,sheet_name="Aging",header=None)
border_grouping=pd.read_excel(file_saving_location,sheet_name="Payer Grouping",header=None)
border_ncr_cal=pd.read_excel(file_saving_location,sheet_name="Payer-Wise NCR Cal",header=None)
border_key=pd.read_excel(file_saving_location,sheet_name="Key",header=None)

interval_border_apply_row=rows_to_keep-1

for i in range(1,no_facility+1+1):
    row_no,col_no=coordinate_finder_insheet(border_ncr_cal,"Facility",i)
    
    border_row=row_no+interval_border_apply_row

    Template=bottom_border_one_cell(Template,"Payer-Wise NCR Cal",border_row,col_no,col_no)

#Raw Data Aging------------------
row_no,col_no=coordinate_finder_insheet(border_aging,"Facility",1)
start_row=row_no+1

for i in range(len(Aging_df)-2):
    Template=left_border(Template,"Aging",start_row+i,col_no,col_no)
    Template=right_border(Template,"Aging",start_row+i,col_no+16,col_no+16)
    
Template=top_border(Template,"Aging",row_no+len(Aging_df)-1,col_no,col_no+16)

#Raw Data Cash------------------
row_no,col_no=coordinate_finder_insheet(border_cash,"Facility",1)
start_row=row_no+1
row_no,col_no_mid=coordinate_finder_insheet(border_cash,"EO month of Apply to Mnth date",1)
row_no,col_no_end=coordinate_finder_insheet(border_cash,"Payer Group - Private Type",1)

for i in range(len(Cash_df)+1):
    Template=left_right_border(Template,"Cash",start_row+i,col_no,col_no)
    Template=left_border(Template,"Cash",start_row+i,col_no_mid,col_no_mid)
    Template=right_border(Template,"Cash",start_row+i,col_no_end,col_no_end)
    
Template=top_border(Template,"Cash",start_row+len(Cash_df)+1,col_no,col_no_end)

#Raw Data AR------------------
row_no,col_no=coordinate_finder_insheet(border_ar,"Facility",1)
start_row=row_no+1
row_no,col_no_mid=coordinate_finder_insheet(border_ar,"EO Month Billing period-Apply date",1)
row_no,col_no_end=coordinate_finder_insheet(border_ar,"Payer Group - Private Type",1)

for i in range(len(AR_df)):
    Template=left_right_border(Template,"AR Detail",start_row+i,col_no,col_no)
    Template=left_border(Template,"AR Detail",start_row+i,col_no_mid,col_no_mid)
    Template=right_border(Template,"AR Detail",start_row+i,col_no_end,col_no_end)
    
Template=top_border(Template,"AR Detail",start_row+len(AR_df),col_no,col_no_end)


#Key-------------------------------------------------
#AR key
row_no,col_no=coordinate_finder_insheet(border_key,"Billing Code",1)
start_row = row_no + 1
for i in range(len(AR_key_df)):
    Template=left_border(Template,"Key",start_row+i,col_no,col_no)
    Template=right_border(Template,"Key",start_row+i,col_no+3,col_no+3)
    
Template=top_border(Template,"Key",start_row+len(AR_key_df),col_no,col_no+3)

#Cash key
row_no,col_no=coordinate_finder_insheet(border_key,"Bill Code",1)
start_row = row_no + 1
for i in range(len(Cash_key_df)):
    Template=left_border(Template,"Key",start_row+i,col_no,col_no)
    Template=right_border(Template,"Key",start_row+i,col_no+1,col_no+1)
    
Template=top_border(Template,"Key",start_row+len(Cash_key_df),col_no,col_no+1)

#Payer Grouping--------------------

#Payer table
row_no,col_no=coordinate_finder_insheet(border_grouping,"Payer",1)
start_row = row_no+1
for i in range(len(payer_group_df)-1):
    Template=left_border(Template,"Payer Grouping",start_row+i,col_no,col_no)
    Template=right_border(Template,"Payer Grouping",start_row+i,col_no+2,col_no+2)
    
Template=top_border(Template,"Payer Grouping",start_row+len(payer_group_df)-1,col_no,col_no+2)

#Types of Payers column border
row_no,col_no=coordinate_finder_insheet(border_grouping,"Payer Groups",1)
for i in range(len(unique_payer_groups)):
    Template=left_right_border(Template,"Payer Grouping",row_no+i,col_no,col_no)
    
Template=left_right_bottom_border(Template,"Payer Grouping",row_no+len(unique_payer_groups),col_no,col_no)

#Types of Payers column border
row_no,col_no=coordinate_finder_insheet(border_grouping,"Payer Group - AR Clean up",1)
for i in range(len(wanted_list)):
    Template=left_right_border(Template,"Payer Grouping",row_no+i,col_no,col_no)
    
Template=left_right_bottom_border(Template,"Payer Grouping",row_no+len(wanted_list),col_no,col_no)

#Facility list column border
row_no,col_no=coordinate_finder_insheet(border_grouping,"Facility List",1)
for i in range(no_facility):
    Template=left_right_border(Template,"Payer Grouping",row_no+i,col_no,col_no)
    
Template=left_right_bottom_border(Template,"Payer Grouping",row_no+no_facility,col_no,col_no)
Template.save(file_saving_location)

#------------------Date Formats--------------------------
app = xw.App(visible=False)
book = xw.Book(file_saving_location)
book.sheets['Cash'].range('Q:Q').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('V:V').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('W:W').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('Y:Y').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('C:C').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('D:D').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('G:G').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('H:H').number_format = 'mm/dd/yyyy'
book.save(file_saving_location)
book.close()
app.kill()

#-------------------------To delete a sheet which not needed (extra generated by Spire)
# Load the Excel file
app = xw.App(visible=False)
wb = xw.Book(file_saving_location)
wb.sheets['Evaluation Warning'].delete()
wb.save(file_saving_location)
wb.close()
app.kill()

print("Assessment Completed")


#Timer
end_time = time.time()
elapsed_time_seconds = end_time - start_time
elapsed_minutes = int(elapsed_time_seconds // 60)
elapsed_seconds = int(elapsed_time_seconds % 60)
print("Elapsed time: ", elapsed_minutes, "minutes", elapsed_seconds, "seconds")

