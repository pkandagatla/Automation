#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
from pandas.tseries.offsets import MonthEnd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import xlwings as xw
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.cell import get_column_letter
import os
import getpass
import sys
import warnings
warnings.filterwarnings("ignore")

#Functions

#Reading raw data
def read_rawdata(filepath):
    app = xw.App(visible=False)
    book = xw.Book(filepath)
    if filepath[-3:] == "xls":
        filepath2 = filepath+"x"
        book.save(filepath2)
        os.remove(filepath)
        df = pd.read_excel(filepath2, index_col=None,header=None)
    else:
        df = pd.read_excel(filepath, index_col=None,header=None)
    book.close()
    app.kill()
    return(df)

def read_data(filepath):
    df = pd.read_excel(filepath, index_col=None,header=None)
    return(df)

#Get the Facility name
def facility_name(df,x,y):
    facility_name=df.iloc[x,y]
    return facility_name

#reordering_column
def reorder_column(df,columnname,position):
    df.insert(position,columnname,df.pop(columnname))
    return df

#To trim a column 
def trim_column(df,column_no):
    df.iloc[:,column_no] = df.iloc[:,column_no].str.strip()
    return df

def trim_many_columns(df,column_nos):
    for j in column_nos:
        for i in range(len(df)):
            if type(df.iloc[i,j]) != str:
                df.iloc[i,j] = str(df.iloc[i,j]).strip()
            else:
                df.iloc[i,j] = df.iloc[i,j].strip()
    return df

def trim_aging(df):
    for i in range(df.shape[0]):
        for j in range(2,df.shape[1]):
            df.iloc[i,j] = float(str(df.iloc[i,j]).replace(" ",'0'))
    return df

#get first row of dataframe
def first_rowofdf(df):
    first_row=pd.DataFrame()
    df=pd.DataFrame(df)
    first_row = df.iloc[0:1, :] 
    return first_row

#It slices the dataframe if it finds a value from "x" space of the dataframe
def iter_row_slicer(df,valuetofind,x):
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                df=df.iloc[i+x:].reset_index(drop=True)
                found=True
                break
        if found is True:
            break
    if(found is False):
        message="No Value"
        return message
    return df


#Here we find a particular values coordinate in a sheet
def coordinate_finder_insheet(template_path,sheetname,valuetofind):
    row_no = 0
    col_no = 0
    df=pd.read_excel(template_path,sheet_name=sheetname,header=None)
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                row_no=i+1
                col_no=j+1
                found=True
                break
        if found is True:
            break 
    return(row_no,col_no)
    
#Here we are pasting the values from a dataframe to the raw data sheet
def copy_paste_raw(Template,sheetname,df,from_row,from_col,to_col):
    
    Sheet=Template[sheetname]

    row_number=from_row
    i=0
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row-1+len(df),min_col=from_col,max_col=to_col):
        j=0
        col_number=from_col
        for cell in row:
            Sheet.cell(row=row_number,column=col_number).value=df.iloc[i,j]
            j+=1
            col_number+=1
        i+=1
        row_number+=1 
    return(Template)

#Here we are extending the formulas till the end we have the values
def extend_formulas(Template,sheetname,df,from_row,from_col,to_col):
    Sheet=Template[sheetname]
    i=from_row
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row-2+len(df),min_col=from_col,max_col=to_col):
        j=from_col
        for cell in row:
            Sheet.cell(row=i,column=j).value=Translator(Sheet.cell(row=from_row-1,column=j).value,origin=Sheet.cell(row=from_row-1,column=j).coordinate).translate_formula(Sheet.cell(row=i,column=j).coordinate)    
            j+=1
        i+=1
    return(Template)


def concatenate_dfs(dfs):
    df = pd.concat(dfs, ignore_index=True)
    return df
    
def first_rowofdf(df):
    first_row=pd.DataFrame()
    df=pd.DataFrame(df)
    first_row = df.iloc[0:1, :] 
    return first_row

def assign_value_to_df(df,x,y,value):
    df.iloc[x,y]=value
    return df

#function to delete values that exist after the particular value
def delete_after_slicer(df,valuetofind,x):   
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                df=df.iloc[0:i+x,:].reset_index(drop=True)  
                found=True
                break
        if found is True:
            break
    return df

#Triggers calculation in the sheet
def perform_calc(filepath):
    wb = xw.Book(filepath)
    app = xw.App(visible=False)
    wb.app.calculate()
    wb.save()
    wb.close()
    app.kill()
    

#Returns a dataframe converted from a excel sheet
def create_df_from_sheet(sheet_name,file_path):
    df=pd.read_excel(file_path,sheet_name=sheet_name)
    return df

#To get a unique value of the dataframe by giving the column indices
def unique_row_values(df,col_no):
    df=df.iloc[:,col_no].unique()
    df=pd.DataFrame(df)
    return df

def text_before(x,character):
    k = []
    for i in x :
        ind = i.find(character)
        k.append(i[0:ind])
    return k

def text_after(x,character):
    k = []
    for i in x :
        ind = i.find(character)
        k.append(i[ind:])
    return k

def unique_values_except(df, omit_list, column_index):
    wanted_list = []
    for i in range(len(df)):
        value = df.iloc[i, column_index]
        if value not in omit_list and value not in wanted_list:
            wanted_list.append(value)
    return wanted_list

#Deletes the below rows in the sheet 
def deleting_below_rows(filepath,sheet_name,start_row,no_of_rows):
    workbook = Workbook()
    workbook.LoadFromFile(filepath)
    worksheet = workbook.Worksheets[sheet_name]
    worksheet.DeleteRow(start_row,no_of_rows)
    workbook.SaveToFile(filepath)
    
#Funtion to paste values
def paste_value_in_cell(workbook,sheet,row_no,col_no,paste_value):
    worksheet=workbook[sheet]
    worksheet.cell(row=row_no,column=col_no).value=paste_value
    return workbook

    
#To delete the specific cells in a range
def delete_cell_range(filepath,sheet_name,start_row,start_col,end_row,end_col):
    if(start_col<end_col ):
        workbook = Workbook()
        workbook.LoadFromFile(filepath)
        worksheet = workbook.Worksheets[sheet_name]   
        range_to_delete = worksheet.Range[start_row,start_col,end_row,end_col] 
        worksheet.DeleteRange(range_to_delete, DeleteOption.MoveLeft)
        workbook.SaveToFile(filepath)
        
def delete_columns(filepath,sheet_name,start_row,start_col,end_row,end_col):
    if(start_col<end_col ):
        workbook = Workbook()
        workbook.LoadFromFile(filepath)
        worksheet = workbook.Worksheets[sheet_name]   
        worksheet.DeleteColumn(start_col,(end_col-start_col))
        workbook.SaveToFile(filepath)
        
#We are iterating to the row to find how many times we have a particular value and if we need it 8 times the 8th found value coordinates is stored
def how_many_values_finder_insheet(template_path,sheetname,valuetofind,no_payers):
    times_recurring=0
    found=False
    df=pd.read_excel(template_path,sheet_name=sheetname,header=None)
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if((df.iloc[i,j]==valuetofind) and (times_recurring<=no_payers)):
                row_no=i+1
                col_no=j+1
                times_recurring+=1
            if(times_recurring==no_payers):
                found=True
                break
        if  found is True:
            break          
    return(row_no,col_no)

#Here we are finding the coordinate with a bound in excel sheet
def coordinate_finder_special(Template,sheetname,value_to_find,from_row,from_col,to_row,to_col):
    
    Sheet=Template[sheetname]
    found=False
    row_number=from_row
    i=0
    for row in Sheet.iter_rows(min_row=from_row,max_row=to_row,min_col=from_col,max_col=to_col):
        j=0
        col_number=from_col
        for cell in row:
            if(Sheet.cell(row=row_number,column=col_number).value==value_to_find):
                found=True
                break
            j+=1
            col_number+=1
        if found is True:
            break
        i+=1
        row_number+=1 
    Template.close()
    return (row_number,col_number)

def to_find_no_months(start_date,end_date):
    dif=relativedelta(end_date,start_date)
    no_months= dif.years*12 + dif.months+1
    return no_months

#Will add a bottom border for a row in a column range
def bottom_border_one_cell(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def right_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_right_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_right_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
    return Template
        

def left_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def right_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def top_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style='medium'), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

#Here we find a particular values coordinate in a sheet with the recurrence at what time
def coordinate_finder_insheet_df(df, valuetofind, recurrence):   
    found = False
    count = 0
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if df.iloc[i, j] == valuetofind:
                row_no = i
                col_no = j
                count += 1
                if count == recurrence:
                    found = True
                    break
        if found:
            break
    if found:
        return (row_no, col_no)
    else:
        return (0, 0) 

import time
start_time = time.time()


# Get the current system's username
username = getpass.getuser()
print("Username:", username)

payer_grp_mail = f"C:/Users/{username}/Downloads/AR Assessment - Payer Grouping.xlsx"

#New user need to store this
path_folder = r"C:/Users/pragna_kandagatla/Desktop/Automation_LPC"

#Checking the files required i.e setup files required and storing
for filename in os.listdir(path_folder):
    file_path4 = os.path.join(path_folder, filename)
    if filename.endswith("AR Assessment - Single Facility Template (LPC).xlsx"):
        template_path = file_path4
    elif filename.endswith("AR Assessment - Payer Grouping.xlsx"):
        payer_grp_send = file_path4

folder_path = sys.argv[1]
folder_path=folder_path.replace("\\","/").replace("\"","")

if __name__ == "__main__":
    folder_path=input("Paste the path of raw data folder= ")
    folder_path=folder_path.replace("\\","/").replace("\"","")

folder_name=folder_path.split("/")[-1]
file_saving_location = f"C:/Users/{username}/Downloads/AR Assessment - {folder_name}.xlsx"

files_read_check = 1
for i, filename in enumerate(os.listdir(folder_path)):
    file_path = os.path.join(folder_path, filename)
    file_path=file_path.replace("\\","/").replace("\"","")
    read_df = read_rawdata(file_path)
    row_no_payer, col_payer = coordinate_finder_insheet_df(read_df, "Payer", 1)
    if read_df.iloc[row_no_payer, col_payer + 1] == "Payer Group":
        payer_group = file_path
        payer_grp_check = "Yes"
        files_read_check += 1
        continue

    read_df = trim_column(read_df,0)
    row_no_aging, col_aging = coordinate_finder_insheet_df(read_df, "Resident Name", 1)
    row_no_ar, col_ar = coordinate_finder_insheet_df(read_df, "Billing Period", 1)
    row_no_cash, col_cash = coordinate_finder_insheet_df(read_df, "GL Account Description", 1)
    
    locals()[f"excel{i}"]  =read_df
    if read_df.iloc[row_no_aging + 1, col_aging] == "Payor Type":
        if file_path[-3:] == "xls":
            file_path = file_path+"x"
        xls_file_aging = file_path
        files_read_check += 1
    elif read_df.iloc[row_no_ar, col_ar+1] == "GL Posting Period":
        if file_path[-3:] == "xls":
            file_path = file_path+"x"
        xls_file_ar = file_path
        files_read_check += 1 
    elif read_df.iloc[row_no_cash + 1, col_cash] == "Resident Name":
        if file_path[-3:] == "xls":
            file_path = file_path+"x"
        xls_file_cash = file_path
        files_read_check += 1
        
if not(files_read_check>4):
    print(f"Some of the raw data files are missing in this folder = {folder_path}")
    sys.exit()
    
# After the loop, you can check the files assigned
print("Files assigned:")
print(f"Cash Report: {xls_file_cash}")
print(f"AR Report: {xls_file_ar}")
print(f"Aging report: {xls_file_aging}")
print(f"Payer group: {payer_group}")

#Loading workbook
Template=load_workbook(filename=template_path)

#----------------------------------Aging-----------------------------------

#Passing Aging report file path to retrieve dataframe of it
Aging_report_df=read_rawdata(xls_file_aging)

#trim first column
Aging_report_df= trim_column(Aging_report_df,0)
Aging_report_df = Aging_report_df.dropna(axis=0, how='all')
Aging_report_df2=Aging_report_df

#Get facility name
A_Facility_Name=facility_name(Aging_report_df,0,0)
Aging_report_df["Facility Name"]=A_Facility_Name

#reorder column
Aging_report_df=reorder_column(Aging_report_df,"Facility Name",0)

#Slice for Resident aging here we need first column to know the aging buckets
Aging_report_df=iter_row_slicer(Aging_report_df,"Resident Name",0)
Aging_report_df_age_bucket =iter_row_slicer(Aging_report_df,"Payor Type",0)

# First row
Month=pd.DataFrame()
Month=first_rowofdf(Aging_report_df)
Age_bucket=pd.DataFrame()
Age_bucket=first_rowofdf(Aging_report_df_age_bucket)

#Initial Raw Data Checks - checking number of columns
check=0
Aging_report_dfheaders=["Facility_Name","Payor Type","Total Due","Current","1 Month","2 Months","3 Months","4 Months","5 Months","6 Months","7 Months","8 Months","9 Months","10 Months","11 Months","12 Months","12+ Months"]

for i in range(1,len(Aging_report_dfheaders)):
    if(Aging_report_dfheaders[i]==Age_bucket.iloc[0,i].strip()):
        check+=1
if(check!=16):
    print(Age_bucket.iloc[0])
    print("Age buckets are in unsual format")

# Checking for Payor Type Summary
aging_check=iter_row_slicer(Aging_report_df2,"Payor Type Summary",1)
if (type(aging_check) == str) | (aging_check.shape[0]<4) | (aging_check.iloc[-1,1] != "Report Totals") :
    print("We don't have Payor Type Summary for the Assessment")

Aging_report_df=iter_row_slicer(Aging_report_df,"Payor Type Summary",2)
Aging_report_df = Aging_report_df.fillna(0)
Aging_report_df = trim_aging(Aging_report_df)

#Concatnate DFs
Header_df=concatenate_dfs([Month,Age_bucket])
Aging_report_df=concatenate_dfs([Header_df,Aging_report_df])
#Give values to some postion in dfs
Aging_report_df=assign_value_to_df(Aging_report_df,0,0,"")
Aging_report_df=assign_value_to_df(Aging_report_df,1,0,"Facility Name")

#function to delete values that exist after the particular value
Aging_report_df=delete_after_slicer(Aging_report_df,"Report Totals",1)
temp = Aging_report_df.iloc[2:,2:].replace(" ","")
temp = temp.replace(np.nan,0)
Aging_report_df.iloc[2:,2:] = temp
print("Aging data loaded")
    
#-------------------------------Cash-------------------------------------------

# Predefined column names
CreportCnames = ["Facility_Name","Resident Name","ID" ,"Trans Date","To Mnth", "Count", "Unit Amount", "Extended Amount", "Bill Code", "Empty1","Private Pay Type","Private Amount","3rd Party Type","3rd Party Amount","Empty2"]
Creportheaders = ["Resident Name","ID" ,"Trans Date","To Mnth", "Count", "Amount", "Amount", "Code", "","Type","Amount","Type","Amount",""]
    
#Passing Cash report file path to retrieve dataframe of it 
Cash_report_df=read_rawdata(xls_file_cash)

#trim first column
Cash_report_df= trim_column(Cash_report_df,0)
Cash_report_df2=Cash_report_df
Cash_report_df2 = Cash_report_df2.fillna("")

#Initial Raw Data checks - checking number of columns
check = 0
for i in range(len(Creportheaders)):
    if(Creportheaders[i]==Cash_report_df2.iloc[6,i].strip()):
        check+=1
if(check!=14):
    print(Cash_report_df2.iloc[6])
    print("Column names does not match for cash report")

C_Facility_Name=facility_name(Cash_report_df,0,0)
Cash_report_df["Facility Name"]=C_Facility_Name

#reorder column
Cash_report_df=reorder_column(Cash_report_df,"Facility Name",0)

#Headers
Cash_report_df = Cash_report_df.set_axis(CreportCnames,axis=1)

#Row slice if it finds the Deposit distribution
Cash_report_df=iter_row_slicer(Cash_report_df,"GL Account Description",4)
    
#function to delete values that exist after the particular value
Cash_report_df=delete_after_slicer(Cash_report_df,"Payor Type Summary",1)
    
#trim columns
Cash_report_df = trim_many_columns(Cash_report_df,[8,10,12])
Cash_report_df = Cash_report_df.replace("nan","")
print("Cash Data Loaded")
    
#-------------------------------------AR Detail-----------------------------------

# Predefined column names
ARreportCnames = ["Billing Period","GL Posting Period" ,"Profile ID","Transaction Source", "Billing Period-Apply", "Transaction Date",                 "Facility Type", "Census Payor Type", "Billing Code","Description","Contractual Allow","Count","Unit Amount",                "Bill Amount","Private Pay Amount", "Private Payor Type","Third Party Amount","Third Party Payor Type",                  "Third Party Bill Type","From Date", "Thru Date"]
                 
#Passing Deposit report file path to retrieve dataframe of it 
AR_report_df=read_rawdata(xls_file_ar)

#Initial Raw Data checks - checking number of columns
check = 0
for i in range(len(ARreportCnames)):
    if(ARreportCnames[i]==AR_report_df.iloc[0,i].strip()):
        check+=1
if(check!=21):
    print(AR_report_df.iloc[0])
    print("Column names does not match for AR report")
    
AR_report_df = AR_report_df.set_axis(ARreportCnames,axis=1)
    
#Get facility name
AR_report_df["Facility Name"]=C_Facility_Name

#reorder column
AR_report_df=reorder_column(AR_report_df,"Facility Name",0)
AR_report_df = trim_many_columns(AR_report_df,[9,10,16,18])
AR_report_df = AR_report_df.replace("nan","")
    
AR_report_df = AR_report_df.iloc[1:,:]
print("AR Data Loaded")
    
#------------------------------------RAW DATA CHECKS---------------------------------

#Facility Name check
if(C_Facility_Name!=A_Facility_Name):
    print("Facility Names are mismatching b/w reports")
    print("  Aging Report= ",A_Facility_Name,"\n  Cash Report= ",C_Facility_Name)
    Aging_report_df["Facility Name"] = C_Facility_Name
        
#Date Range Check
a_start_index=pd.to_datetime(AR_report_df.iloc[0,1])
a_end_index=pd.to_datetime(AR_report_df.iloc[-1,1])
a_date_range = str(a_start_index)[:10] + " - " + str(a_end_index)[:10]

c_start_index = pd.to_datetime(Cash_report_df2.iloc[2,0].split(" ")[7]) + MonthEnd(0)
c_end_index = pd.to_datetime(Cash_report_df2.iloc[2,0].split(" ")[9]) + MonthEnd(0)
c_date_range = str(c_start_index)[:10] + " - " + str(c_end_index)[:10]

aging_end_index = pd.to_datetime(Aging_report_df2.iloc[1,4].strip().split(" ")[-1])

if((a_start_index == c_start_index) & (a_end_index == c_end_index == aging_end_index)):
    print("Date Ranges are matching")
else:
    print("We have different Date Ranges for the reports \n  Cash Report= ",c_date_range,"\n  AR Detail Report= ",a_date_range,"\n  Aging Report= ",str(aging_end_index)[:10])

print("Performed all raw data checks!")

#--------------------------------------Payers Extraction---------------------------------------------------

mask = Aging_report_df[0] == 0
Aging_report_df = Aging_report_df[~mask]
Aging_payers = Aging_report_df[0]
Aging_payers_df = Aging_payers.dropna()

Cash_payers_priv = Cash_report_df['Private Pay Type']
Cash_payers_third = Cash_report_df['3rd Party Type']
AR_payers_priv = AR_report_df['Private Payor Type']
AR_payers_third = AR_report_df['Third Party Payor Type']

payers_df = concatenate_dfs([Cash_payers_priv,Cash_payers_third,AR_payers_priv,AR_payers_third])
payers_df = payers_df.dropna()

unique_payers_ncr = payers_df.unique()
unique_payers_ncr  = list(filter(lambda x: x.find('MAGIC') == -1 , unique_payers_ncr))
unique_payers_ncr=list(filter(lambda x: len(x)>0, unique_payers_ncr))

Aging_payers_df = Aging_payers_df.unique()
Aging_payers_df  = list(filter(lambda x: x.find('Resident Name') == -1 , Aging_payers_df))
Aging_payers_df  = list(filter(lambda x: x.find('Payor Type') == -1 , Aging_payers_df))
Aging_payers_df  = list(filter(lambda x: x.find('Report Totals') == -1 , Aging_payers_df))
Aging_payers_df=list(filter(lambda x: len(x)>0, Aging_payers_df))

Aging_payers_payortype = text_before(Aging_payers_df," ")
Aging_payers_payor = Aging_payers_df
Aging_payers_final = pd.DataFrame({'Payor Type' : Aging_payers_payortype,
                                'Payor' : Aging_payers_payor })

Aging_payers_final = trim_column(Aging_payers_final,1)
ncr_payers_df = pd.DataFrame({'Payor Type' : unique_payers_ncr})

payers_list = pd.merge(Aging_payers_final, ncr_payers_df, on='Payor Type', how='outer')

#Pasting the raw data in the Payer group template
payer_grp_file=load_workbook(payer_grp_send)

#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(payer_grp_send,"Payer Grouping","Payer")

#Paste values in excel
payer_grp_file=copy_paste_raw(payer_grp_file,"Payer Grouping",payers_list,row_no+1,col_no,col_no+1)

#Saving the file
payer_grp_file.save(payer_grp_mail)
perform_calc(payer_grp_mail)

print("Payer Group file is Generated!")
#--------------------------------Payer Group file upload------------------------------------------
if(payer_grp_check=="Yes"):
        
    #Payer Group Dataframe
    payer_group_df= read_data(payer_group)
    payer_group_df.columns=payer_group_df.iloc[1]

    #Slice for Payer
    payer_group_df=iter_row_slicer(payer_group_df,"Payer",1)
    payer_group_df=payer_group_df.iloc[:,1:4]
    
    payer_group_df_paste=payer_group_df.iloc[:,:3]

    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_path,"Payer Grouping","Payer")

    #Paste values in excel
    Template=copy_paste_raw(Template,"Payer Grouping",payer_group_df_paste,row_no+1,col_no,col_no+2)
    
    #Payer Grouping copy paste
    row_no, col_no=coordinate_finder_insheet(template_path,"Payer Grouping","Payer Groups")
    
    #Getting value for number of payers in the assessment
    payer_group_df.loc[-1] = ["", "", "Private"]  
    payer_group_df.index = payer_group_df.index + 1 
    payer_group_df = payer_group_df.sort_index()
    payer_group_df=trim_column(payer_group_df,2)
    unique_payer_groups=unique_values_except(payer_group_df,["Non-QHCR Billed Payer"],2)
    unique_payer_groups = pd.DataFrame(unique_payer_groups)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Payer Grouping",unique_payer_groups,row_no+1,col_no,col_no)
    
    
print("Payer Grouping tab updated")



#------------------------------------Pasting Raw Data in Excel-----------------------------------

#----------------------CashReport-------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_path,"Cash","Facility Name")
to_row,to_col=coordinate_finder_insheet(template_path,"Cash","Empty#")

#Paste values in excel
Template =copy_paste_raw(Template,"Cash",Cash_report_df,row_no+1,col_no,to_col)

#Formula pasting range finder
for_row_no,for_col_no=coordinate_finder_insheet(template_path,"Cash","EO month of Apply to Mnth date")
for_to_row,for_to_col=coordinate_finder_insheet(template_path,"Cash","Payer Group - Private Type")

#Extend Formulas in Sheets
Template=extend_formulas(Template,"Cash",Cash_report_df,for_row_no+2,for_col_no,for_to_col)


#------------------------------AR Detail Report---------------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_path,"AR Detail","Facility Name")
to_row,to_col=coordinate_finder_insheet(template_path,"AR Detail","Thru Date")

#Paste values in excel
Template =copy_paste_raw(Template,"AR Detail",AR_report_df,row_no+1,col_no,to_col)

#Formula pasting range finder
for_row_no,for_col_no=coordinate_finder_insheet(template_path,"AR Detail","EO Month Billing period-Apply date")
for_to_row,for_to_col=coordinate_finder_insheet(template_path,"AR Detail","Payer Group - Private Type")

#Extend Formulas in Sheets
Template=extend_formulas(Template,"AR Detail",AR_report_df,for_row_no+2,for_col_no,for_to_col)


#----------------------------------------Aging----------------------------------------

#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_path,"Aging","Facility Name")
to_row,to_col=coordinate_finder_insheet(template_path,"Aging","12+ Months")

#Paste values in excel
Template=copy_paste_raw(Template,"Aging",Aging_report_df,row_no-1,col_no,to_col)


#------------------------------------------Cash Key---------------------------------------
Cash_key = Cash_report_df['Bill Code'].unique()
Cash_key_df = pd.DataFrame({"Bill Code":Cash_key})
Cash_key_df['Bill Code'] = Cash_key_df['Bill Code'].replace('', np.nan)
Cash_key_df = Cash_key_df.dropna()
Cash_key_df['Include'] = [1 if x == 'PAY' else 0 for x in Cash_key_df['Bill Code']]

#Cash key copy paste
row_no, col_no=coordinate_finder_insheet(template_path,"Key","Bill Code")

Template=copy_paste_raw(Template,"Key",Cash_key_df,row_no+1,col_no,col_no+1)

#-------------------------------------AR key---------------------------------------------------
AR_key_df = AR_report_df.loc[:,['Billing Code','Description']]
AR_key_df = AR_key_df.drop_duplicates()

AR_key_df["Include"] = 1
AR_key_df["Bad debt"] = 0
AR_key_df = AR_key_df.reset_index()

for i in range(len(AR_key_df)):
    if AR_key_df["Billing Code"][i] == "PAY":
        AR_key_df["Include"][i] = 0
    elif (AR_key_df["Description"][i].lower()).find("refund") != -1:
        AR_key_df["Include"][i] = 2
    elif (AR_key_df["Description"][i].lower()).find("bad debt") != -1:
        AR_key_df["Include"][i] = 3
        AR_key_df["Bad debt"][i] = 1
    elif ((AR_key_df["Description"][i].lower()).find("c/a") != -1) or ((AR_key_df["Description"][i].lower()).find("c/adj") != -1) or ((AR_key_df["Description"][i].lower()).find("contractual adjustment") != -1):
        AR_key_df["Include"][i] = 3
        AR_key_df["Bad debt"][i] = 2
    elif ((AR_key_df["Description"][i].lower()).find("w/o") != -1) or ((AR_key_df["Description"][i].lower()).find("write-off") != -1) or ((AR_key_df["Description"][i].lower()).find("write off") != -1):
        AR_key_df["Include"][i] = 3
        
AR_key_df = AR_key_df.drop("index",axis=1)
        
#AR key copy paste
row_no, col_no=coordinate_finder_insheet(template_path,"Key","Billing Code")

Template=copy_paste_raw(Template,"Key",AR_key_df,row_no+1,col_no,col_no+3)

print("Key tab updated")

#------------------------------------------Calculation Sheets--------------------------------------
#----------------------------------------------NCR-----------------------------------
#to get number of months
no_months=to_find_no_months(a_start_index,a_end_index)

#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_path,"NCR","Metrics")

#To paste the end range month for assessment period
Template=paste_value_in_cell(Template,"NCR",row_no,col_no+19,a_end_index)

#-----------------------------------------AR Clean Up Estimates------------------------------------------

#Unique list of payers need to be added in the AR Clean up estimates
wanted_list = unique_values_except(payer_group_df, ["Private", "Other", "Miscellaneous","Non-QHCR Billed Payer"], 2)
wanted_list=pd.DataFrame(wanted_list)

#Getting position
row_num, col_num=coordinate_finder_insheet(template_path,"AR Clean Up estimates","Payer Type")
row_end, col_end=coordinate_finder_insheet(template_path,"AR Clean Up estimates","Payer Type Total")

#Paste values in excel
Template=copy_paste_raw(Template,"AR Clean Up estimates",wanted_list,row_num+1,col_num,col_num)

#-----------------------------------------Aging Cal------------------------------------------

#Unique list of payer type need to be added in the Aging Cal
unique_payers_type = unique_values_except(payer_group_df,["Non-QHCR Billed Payer"],2)
unique_payers_type = pd.DataFrame(unique_payers_type)

#Getting position
row_num, col_num=coordinate_finder_insheet(template_path,"Aging Cal","Payer Group Summary")
row_end, col_end=coordinate_finder_insheet(template_path,"Aging Cal","Payer Type Total")

#Paste values in excel
Template=copy_paste_raw(Template,"Aging Cal",unique_payers_type,row_num+1,col_num,col_num)

#Unique list of payers need to be added in the Aging Cal
unique_payers = unique_values_except(payer_group_df,["",np.nan],1)
unique_payers = pd.DataFrame(unique_payers)

#Getting position
row_num, col_num=coordinate_finder_insheet(template_path,"Aging Cal","Payer Type Summary")
row_end, col_end=coordinate_finder_insheet(template_path,"Aging Cal","Payer Type Total")

#Paste values in excel
Template=copy_paste_raw(Template,"Aging Cal",unique_payers,row_num+1,col_num,col_num)


print("Values needed to be pasted in Calculation sheets are done!")
print("Formatting of cells started......")

#-------------------------------------------- Formatting raw data sheets

#Raw Data Aging------------------
row_no,col_no=coordinate_finder_insheet(template_path,"Aging","Facility Name")
start_row=row_no+1

for i in range(len(Aging_report_df)-2):
    Template=left_border(Template,"Aging",start_row+i,col_no,col_no)
    Template=right_border(Template,"Aging",start_row+i,col_no+16,col_no+16)
    
Template=top_border(Template,"Aging",row_no+len(Aging_report_df)-1,col_no,col_no+16)


#Raw Data Cash------------------
row_no,col_no=coordinate_finder_insheet(template_path,"Cash","Facility Name")
start_row=row_no+1
row_no,col_no_mid=coordinate_finder_insheet(template_path,"Cash","EO month of Apply to Mnth date")
row_no,col_no_end=coordinate_finder_insheet(template_path,"Cash","Payer Group - Private Type")

for i in range(len(Cash_report_df)):
    Template=left_right_border(Template,"Cash",start_row+i,col_no,col_no)
    Template=left_border(Template,"Cash",start_row+i,col_no_mid,col_no_mid)
    Template=right_border(Template,"Cash",start_row+i,col_no_end,col_no_end)
    
Template=top_border(Template,"Cash",start_row+len(Cash_report_df),col_no,col_no_end)

#Raw Data AR------------------
row_no,col_no=coordinate_finder_insheet(template_path,"AR Detail","Facility Name")
start_row=row_no+1
row_no,col_no_mid=coordinate_finder_insheet(template_path,"AR Detail","EO Month Billing period-Apply date")
row_no,col_no_end=coordinate_finder_insheet(template_path,"AR Detail","Payer Group - Private Type")

for i in range(len(AR_report_df)):
    Template=left_right_border(Template,"AR Detail",start_row+i,col_no,col_no)
    Template=left_border(Template,"AR Detail",start_row+i,col_no_mid,col_no_mid)
    Template=right_border(Template,"AR Detail",start_row+i,col_no_end,col_no_end)
    
Template=top_border(Template,"AR Detail",start_row+len(AR_report_df),col_no,col_no_end)


#Key-------------------------------------------------
#AR key
row_no,col_no=coordinate_finder_insheet(template_path,"Key","Billing Code")
start_row = row_no + 1
for i in range(len(AR_key_df)):
    Template=left_border(Template,"Key",start_row+i,col_no,col_no)
    Template=right_border(Template,"Key",start_row+i,col_no+3,col_no+3)
    
Template=top_border(Template,"Key",start_row+len(AR_key_df),col_no,col_no+3)

#Cash key
row_no,col_no=coordinate_finder_insheet(template_path,"Key","Bill Code")
start_row = row_no + 1
for i in range(len(Cash_key_df)):
    Template=left_border(Template,"Key",start_row+i,col_no,col_no)
    Template=right_border(Template,"Key",start_row+i,col_no+1,col_no+1)
    
Template=top_border(Template,"Key",start_row+len(Cash_key_df),col_no,col_no+1)


#Payer Grouping--------------------

#Payer table
row_no,col_no=coordinate_finder_insheet(template_path,"Payer Grouping","Payer")
start_row = row_no+1
for i in range(len(payer_group_df)-1):
    Template=left_border(Template,"Payer Grouping",start_row+i,col_no,col_no)
    Template=right_border(Template,"Payer Grouping",start_row+i,col_no+2,col_no+2)
    
Template=top_border(Template,"Payer Grouping",start_row+len(payer_group_df)-1,col_no,col_no+2)


#Types of Payers column border
row_no,col_no=coordinate_finder_insheet(template_path,"Payer Grouping","Payer Groups")
for i in range(len(unique_payer_groups)):
    Template=left_right_border(Template,"Payer Grouping",row_no+i,col_no,col_no)
    
Template=left_right_bottom_border(Template,"Payer Grouping",row_no+len(unique_payer_groups),col_no,col_no)
Template.save(file_saving_location)

#------------------Date Formats--------------------------
app = xw.App(visible=False)
book = xw.Book(file_saving_location)
book.sheets['Cash'].range('Q:Q').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('Y:Y').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('V:V').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('W:W').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('C:C').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('D:D').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('G:G').number_format = 'mm/dd/yyyy'
book.sheets['AR Detail'].range('H:H').number_format = 'mm/dd/yyyy'
book.save(file_saving_location)
book.close()
app.kill()

print("Deletion of rows/cells started......")

from spire.xls.common import *
from spire.xls import *

#--------------------------------Deletion of row starts-----------------------------

#---------------------------------------------AR Clean up Estimates----------------------------------------------
#Getting position
row_num, col_num=coordinate_finder_insheet(template_path,"AR Clean Up estimates","Payer Type")
row_end, col_end=coordinate_finder_insheet(template_path,"AR Clean Up estimates","Payer Type Total")

#Calculation to find the start row nd number of rows to delete
start_row=row_num+len(wanted_list)+1
no_of_row=row_end-start_row

#Delete unwated rows
deleting_below_rows(file_saving_location,"AR Clean Up estimates",start_row,no_of_row)

#--------------------------------Aging Cal--------------------------------------------

#Getting position
row_no,col_no=coordinate_finder_insheet(template_path,"Aging Cal","Payer Group Summary")
row_end,col_end=coordinate_finder_insheet(template_path,"Aging Cal","Payer Type Total")

#Calculation to find the start row nd number of rows to delete
start_row=row_no+len(unique_payers_type)+1
no_of_row=row_end-start_row

#Delete unwated rows
deleting_below_rows(file_saving_location,"Aging Cal",start_row,no_of_row)

app = xw.App(visible=False)
book = xw.Book(file_saving_location)
book.save(file_saving_location)
book.close()
app.kill()

#Getting Position - payers
row_no,col_no=coordinate_finder_insheet(file_saving_location,"Aging Cal","Payer Type Summary")
row_end,col_end=coordinate_finder_insheet(file_saving_location,"Aging Cal","Payer Total")


#Delete unwated rows in Payers
start_row = row_no+len(unique_payers)+1
no_of_row = row_end-start_row

deleting_below_rows(file_saving_location,"Aging Cal",start_row,no_of_row)


#--------------------------------------NCR----------------------------------------------------
#Find the position of the particular excel paths
row_no,col_no=coordinate_finder_insheet(template_path,"NCR","Metrics")

delete_columns(file_saving_location,"NCR",row_no-2,col_no+2,10000,col_no+18-no_months)

#Find the position of the particular value for the nth time where the last payers table is found
last_metrics_row,last_metrics_col=how_many_values_finder_insheet(template_path,"NCR","Metrics",len(unique_payer_groups)+2)

# Load the Excel workbook
Template = load_workbook(filename=template_path)
#Here finding the empty cell and then we can delete from there
row_no,col_no=coordinate_finder_special(Template,"NCR",None,last_metrics_row,last_metrics_col,10000,last_metrics_col+no_months+2)

deleting_below_rows(file_saving_location,"NCR",row_no,10000)



#-------------------------To delete a sheet which not needed (extra generated by Spire)
# Load the Excel file

app = xw.App(visible=False)
wb = xw.Book(file_saving_location)
wb.sheets['Evaluation Warning'].delete()
wb.save(file_saving_location)
wb.close()
app.kill()

print("Assessment Completed")


#Timer
end_time = time.time()
elapsed_time_seconds = end_time - start_time
elapsed_minutes = int(elapsed_time_seconds // 60)
elapsed_seconds = int(elapsed_time_seconds % 60)
print("Elapsed time: ", elapsed_minutes, "minutes", elapsed_seconds, "seconds")

