#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import pandas as pd
import numpy as np
from pandas.tseries.offsets import MonthEnd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import xlwings as xw
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.cell import get_column_letter
import getpass
import sys
from copy import copy
from spire.xls.common import *
from spire.xls import *
import warnings
warnings.filterwarnings("ignore")
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList

#Functions

#Reading raw data    
def read_data(filepath,month):
    df = pd.read_excel(filepath, index_col=None)
    df.iloc[0,:] = df.iloc[0,:].fillna("Unknown")
    df["Month"] = month
    df = reorder_column(df,"Month",0)
    return(df)

def read_raw_data(filepath):
    df = pd.read_excel(filepath, index_col=None, header=None)
    return(df)

def reorder_column(df,columnname,position):
    df.insert(position,columnname,df.pop(columnname))
    return df

def concatenate_dfs(dfs):
    df = pd.concat(dfs, ignore_index=True)
    return df

def read_ar_data(ar_path,month):
    df = read_data(ar_path,month)
    row_no, col_no = coordinate_finder_insheet_df(df,"Patient",1)
    df.columns = df.iloc[row_no,:]
    df = df.iloc[row_no+1:,:]
    if np.nan in df.columns:
        df = df.drop(np.nan,axis=1)
    return df

def concatenate_ar_data(ar_ins_path,ar_pat_path,month):
    df1 = read_ar_data(ar_ins_path,month)
    df2 = read_ar_data(ar_pat_path,month)
    df = concatenate_dfs([df1,df2])
    return df

#Here we find a particular values coordinate in a sheet with the recurrence at what time
def coordinate_finder_insheet_df(df, valuetofind, recurrence):   
    found = False
    count = 0
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if df.iloc[i, j] == valuetofind:
                row_no = i
                col_no = j
                count += 1
                if count == recurrence:
                    found = True
                    break
        if found:
            break
    if found:
        return (row_no, col_no)
    else:
        return (0, 0) 

def check_raw_data(path):
    files_read_check = 1
    ncr_path = ""
    ar_ins_path = ""
    ar_pat_path = ""
    fin_path = ""
    
    for j, filename in enumerate(os.listdir(path)):
        file_path = os.path.join(path, filename)
        file_path=file_path.replace("\\","/").replace("\"","")
        read_df = read_raw_data(file_path)
        row_no_ncr, col_ncr = coordinate_finder_insheet_df(read_df, "Remaining Balance", 1)
        row_no_ar_ins, col_ar_ins = coordinate_finder_insheet_df(read_df, "Responsibilities: Primary, Secondary, Tertiary", 1)
        row_no_ar_pat, col_ar_pat = coordinate_finder_insheet_df(read_df, "Responsibilities: Patient", 1)
        row_no_fin, col_fin = coordinate_finder_insheet_df(read_df, "Charges", 1)

        if read_df.iloc[row_no_ncr, col_ncr+1] == "Net Collection Rate":
            ncr_path = file_path
            files_read_check += 1
        elif read_df.iloc[row_no_ar_ins+8, col_ar_ins] == "Patient":
            ar_ins_path = file_path
            files_read_check += 1 
        elif read_df.iloc[row_no_ar_pat+8, col_ar_pat] == "Patient":
            ar_pat_path = file_path
            files_read_check += 1
        elif read_df.iloc[row_no_fin, col_fin+1] == "Total Adjustments":
            fin_path = file_path
            files_read_check += 1
    return files_read_check, ncr_path, ar_ins_path, ar_pat_path, fin_path

def copy_paste_raw(Template,sheetname,df,from_row,from_col,to_col):
    
    Sheet=Template[sheetname]

    row_number=from_row
    i=0
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row-1+len(df),min_col=from_col,max_col=to_col):
        j=0
        col_number=from_col
        for cell in row:
            Sheet.cell(row=row_number,column=col_number).value=df.iloc[i,j]
            j+=1
            col_number+=1
        i+=1
        row_number+=1 
    return(Template)

def paste_value_in_cell(workbook,sheet,row_no,col_no,paste_value):
    worksheet=workbook[sheet]
    worksheet.cell(row=row_no,column=col_no).value=paste_value
    return workbook

def format_cells(Template,sheetname,row_no_start,row_no_end,col_no_start,col_no_end):
    sheet=Template[sheetname]
    for row in range(row_no_start, row_no_end+1):
        for col_no in range(col_no_start,col_no_end+1):
            cell = sheet.cell(row=row, column=col_no)
            cell.border = copy(sheet.cell(row=4, column=col_no).border)
            cell.fill = copy(sheet.cell(row=4, column=col_no).fill)
            cell.number_format = copy(sheet.cell(row=4, column=col_no).number_format)
            cell.font = copy(sheet.cell(row=4, column=col_no).font)
            cell.alignment = copy(sheet.cell(row=4, column=col_no).alignment)
        
    return Template

def bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_right_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_border_one_cell(Template,sheetname,row_no,col_no):
    sheet=Template[sheetname]
    border = Border(left=Side(style='medium'), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))
    
    cell = sheet.cell(row=row_no, column=col_no)
    cell.border = border
        
    return Template

def right_border_one_cell(Template,sheetname,row_no,col_no):
    sheet=Template[sheetname]
    border = Border(left=Side(style=None), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))
    
    cell = sheet.cell(row=row_no, column=col_no)
    cell.border = border
        
    return Template

def left_right_border(Template,sheetname,row_start,row_end,col_no):
    sheet=Template[sheetname]
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for row in range(row_start, row_end+1):
        cell = sheet.cell(row=row, column=col_no)
        cell.border = border
        
    return Template

def insert_row_ar(workbook,template,worksheet,number,template_providers,new_providers_ar):
    row_no, col_no = coordinate_finder_insheet_df(template,"Parameter",2)
    worksheet = workbook.Worksheets[worksheet]
    worksheet.InsertRow(row_no + 2 + len(template_providers)*number-number, len(new_providers_ar)*number)
    
def formula(Template,sheetname,row_start,row_end,col_start,col_end,number):
    ws = Template[sheetname]
    for row in range(row_start,row_end+1):
        for col in range(col_start,col_end+1):   
            source_cell = ws.cell(row=row-number, column=col)
            dest_cell = ws.cell(row=row, column=col)
            formula = source_cell.value
            translated = Translator(formula, source_cell.coordinate).translate_formula(dest_cell.coordinate)
            ws[dest_cell.coordinate] = translated
    return Template

def formula_merged_cell(Template,sheetname,row,col,number):
    ws = Template[sheetname] 
    source_cell = ws.cell(row=row-number, column=col)
    dest_cell = ws.cell(row=row, column=col)
    formula = source_cell.value
    translated = Translator(formula, source_cell.coordinate).translate_formula(dest_cell.coordinate)
    ws[dest_cell.coordinate] = translated
    return Template

def formula_provider(Template,sheetname,row,col,number):
    ws = Template[sheetname] 
    source_cell = ws.cell(row=row-number, column=col)
    dest_cell = ws.cell(row=row, column=col)
    formula = source_cell.value
    translated = Translator(formula, source_cell.coordinate).translate_formula(dest_cell.coordinate)
    position = translated.find("AE$") + 3
    n = int(translated[position:])+1
    translated = translated[:position] + str(n)
    ws[dest_cell.coordinate] = translated
    return Template

def format_cells_ar(Template,sheetname,row_no_start,row_no_end,col_no_start,col_no_end,number):
    sheet=Template[sheetname]
    for row in range(row_no_start, row_no_end):
        for col_no in range(col_no_start+1,col_no_end+2):
            cell = sheet.cell(row=row, column=col_no)
            cell.border = copy(sheet.cell(row=row-number, column=col_no).border)
            cell.fill = copy(sheet.cell(row=row-number, column=col_no).fill)
            cell.number_format = copy(sheet.cell(row=row-number, column=col_no).number_format)
            cell.font = copy(sheet.cell(row=row-number, column=col_no).font)
            cell.alignment = copy(sheet.cell(row=row-number, column=col_no).alignment)
        
    return Template

def formatting_ar_sheets(template_providers,new_providers_ar,all_providers_ar,file_saving_location,sheetname,template_ar,number):
    
    Template = load_workbook(file_saving_location)

    row_no, col_no_start = coordinate_finder_insheet_df(template_ar,"Provider",2)
    row_no_new, col_no_end = coordinate_finder_insheet_df(template_ar,"Total",1)

    start_row = row_no + 2 + (len(template_providers)*number)-number
    no_of_rows = (len(new_providers_ar))*number
    end_row = start_row  + no_of_rows
    
    Template = format_cells_ar(Template,sheetname,start_row,end_row,col_no_start,col_no_end,number)
    Template = formula(Template,sheetname,start_row,end_row,col_no_start+3,col_no_end+1,number)
                               
    cell_range = range(start_row,end_row-1,number)
    ws=Template[sheetname]

    for i in cell_range:
        ws.merge_cells(start_row=i, start_column=col_no_start+1, end_row=i+(number-1), end_column=col_no_start+1)
        ws.merge_cells(start_row=i, start_column=col_no_start+2, end_row=i+(number-1), end_column=col_no_start+2)

    cell_range = range(start_row,end_row-1+number,number)

    for i in cell_range:
        Template = formula_merged_cell(Template,sheetname,i,col_no_start+2,number)
        Template = formula_provider(Template,sheetname,i,col_no_start+1,number)


    cell_range = range(start_row+2,end_row+3-1,3)

    row_no, col_no = coordinate_finder_insheet_df(template_ar,"0-30",1)

    for i in cell_range:
        ws.merge_cells(start_row=i, start_column=col_no+1, end_row=i, end_column=col_no+3)

    ws.conditional_formatting = ConditionalFormattingList()

    redFill = PatternFill(start_color='FF5E6C',end_color='FF5E6C',fill_type='solid')
    greenFill = PatternFill(start_color='A9D08E',end_color='A9D08E',fill_type='solid')

    start_row = row_no+4
    end_row = start_row+6*number
    cell_range = range(start_row,end_row,3)

    for i in cell_range:
        ws.conditional_formatting.add('E'+str(i), CellIsRule(operator='lessThan', formula=[0.7], stopIfTrue=True, fill=redFill))
        ws.conditional_formatting.add('E'+str(i), CellIsRule(operator='greaterThanOrEqual', formula=[0.7], stopIfTrue=True, fill=greenFill))
        ws.conditional_formatting.add('H'+str(i), CellIsRule(operator='greaterThanOrEqual', formula=[0.2], stopIfTrue=True, fill=redFill))
        ws.conditional_formatting.add('H'+str(i), CellIsRule(operator='lessThan', formula=[0.2], stopIfTrue=True, fill=greenFill))
        ws.conditional_formatting.add('I'+str(i), CellIsRule(operator='greaterThanOrEqual', formula=[0.1], stopIfTrue=True, fill=redFill))
        ws.conditional_formatting.add('I'+str(i), CellIsRule(operator='lessThan', formula=[0.1], stopIfTrue=True, fill=greenFill))

    row_no, col_no = coordinate_finder_insheet_df(template_ar,"0-30",2)

    start_row = row_no+4
    end_row = start_row+(len(all_providers_ar)+1)*number
    cell_range = range(start_row,end_row,3)

    for i in cell_range:
        ws.conditional_formatting.add('E'+str(i), CellIsRule(operator='lessThan', formula=[0.7], stopIfTrue=True, fill=redFill))
        ws.conditional_formatting.add('E'+str(i), CellIsRule(operator='greaterThanOrEqual', formula=[0.7], stopIfTrue=True, fill=greenFill))
        ws.conditional_formatting.add('H'+str(i), CellIsRule(operator='greaterThanOrEqual', formula=[0.2], stopIfTrue=True, fill=redFill))
        ws.conditional_formatting.add('H'+str(i), CellIsRule(operator='lessThan', formula=[0.2], stopIfTrue=True, fill=greenFill))
        ws.conditional_formatting.add('I'+str(i), CellIsRule(operator='greaterThanOrEqual', formula=[0.1], stopIfTrue=True, fill=redFill))
        ws.conditional_formatting.add('I'+str(i), CellIsRule(operator='lessThan', formula=[0.1], stopIfTrue=True, fill=greenFill))

    Template.save(file_saving_location)
    
    
def execute(template_path,file_saving_location,fin_path,ncr_path,ar_ins_path,ar_pat_path,current_month):
    #Loading workbook
    Template=load_workbook(filename=template_path)

    #storing all the sheets as a dataframe
    template_ar_worklist=pd.read_excel(template_path,sheet_name="AR WorkList",header=None)
    template_fin_summary=pd.read_excel(template_path,sheet_name="Financial Activity Report",header=None)
    template_charges_summary=pd.read_excel(template_path,sheet_name="Charges Collected Summary",header=None)
    template_ncr=pd.read_excel(template_path,sheet_name="NCR",header=None)
    template_ar=pd.read_excel(template_path,sheet_name="AR",header=None)
    template_ins_ar=pd.read_excel(template_path,sheet_name="Insurance AR",header=None)
    template_pat_ar=pd.read_excel(template_path,sheet_name="Patient AR",header=None)
    template_summary=pd.read_excel(template_path,sheet_name="Executive Summary",header=None)

    print("Template files are converted to dataframe")

    #---------------------- Reading data

    AR_report_df = concatenate_ar_data(ar_ins_path,ar_pat_path,current_month)
    Fin_report_df = read_data(fin_path,current_month)
    if ncr_path != '':
        Ncr_report_df = read_data(ncr_path,current_month)
    print("Raw Data Files are loaded")

    #------------------------------------Pasting Raw Data and formatting in Excel-----------------------------------

    #----------------------AR Worklist Report-------------------------------
    row_no,col_no=coordinate_finder_insheet_df(template_ar_worklist,"Snapshot Date",1)
    to_row,to_col=coordinate_finder_insheet_df(template_ar_worklist,"Last Worklist Status Username",1)

    #Paste values in excel
    Template =copy_paste_raw(Template,"AR WorkList",AR_report_df,len(template_ar_worklist)+1,col_no+1,to_col+1)
    Template = format_cells(Template,"AR WorkList",len(template_ar_worklist),len(template_ar_worklist)+len(AR_report_df),col_no+1,to_col+1)
    Template = bottom_border(Template,"AR WorkList", len(template_ar_worklist)+len(AR_report_df),col_no+1,to_col+1)
    Template = left_border_one_cell(Template,"AR WorkList", len(template_ar_worklist)+len(AR_report_df),col_no+1)
    Template = right_border_one_cell(Template,"AR WorkList", len(template_ar_worklist)+len(AR_report_df),to_col+1)



    #---------------------Financial Activity Report-------------------------------
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet_df(template_fin_summary,"Month",1)
    to_row,to_col=coordinate_finder_insheet_df(template_fin_summary,"Total Applied Payments",1)

    #Paste values in excel
    Template =copy_paste_raw(Template,"Financial Activity Report",Fin_report_df,len(template_fin_summary)+1,col_no+1,to_col+1)
    Template = format_cells(Template,"Financial Activity Report",len(template_fin_summary),len(template_fin_summary)+len(Fin_report_df),col_no+1,to_col+1)
    Template = bottom_border(Template,"Financial Activity Report", len(template_fin_summary)+len(Fin_report_df),col_no+1,to_col+1)
    Template = left_border_one_cell(Template,"Financial Activity Report", len(template_fin_summary)+len(Fin_report_df),col_no+1)
    Template = right_border_one_cell(Template,"Financial Activity Report", len(template_fin_summary)+len(Fin_report_df),to_col+1)



    #---------------------Charges Collected Summary Report-------------------------------
    #Find the position of the particular excel paths
    if ncr_path != '':
        row_no,col_no=coordinate_finder_insheet_df(template_charges_summary,"Month",1)
        to_row,to_col=coordinate_finder_insheet_df(template_charges_summary,"Net Collection Rate",1)

        #Paste values in excel
        Template =copy_paste_raw(Template,"Charges Collected Summary",Ncr_report_df,len(template_charges_summary)+1,col_no+1,to_col+1)
        Template = format_cells(Template,"Charges Collected Summary",len(template_charges_summary),len(template_charges_summary)+len(Ncr_report_df),col_no+1,to_col+1)
        Template = bottom_border(Template,"Charges Collected Summary", len(template_charges_summary)+len(Ncr_report_df),col_no+1,to_col+1)
        Template = left_border_one_cell(Template,"Charges Collected Summary", len(template_charges_summary)+len(Ncr_report_df),col_no+1)
        Template = right_border_one_cell(Template,"Charges Collected Summary", len(template_charges_summary)+len(Ncr_report_df),to_col+1)

    #Paste current month
    row_no,col_no=coordinate_finder_insheet_df(template_summary,"Current Month",1)
    Template = paste_value_in_cell(Template,"Executive Summary",row_no+1,col_no+2,current_month)

    print("Pasting values and Formatting raw data tabs done")

    row_no, col_no = coordinate_finder_insheet_df(template_ncr,"Provider",2)
    Providers_template_ncr = list(template_ncr.iloc[row_no+1:,1])
    new_providers = []

    if ncr_path != '':
        Providers_current_month_ncr = list(Ncr_report_df.iloc[:,1])

        for i in Providers_current_month_ncr:
            if i in Providers_template_ncr or i == "Report Totals":
                continue
            else:
                new_providers.append(i)

    Providers_current_month_fin = list(Fin_report_df.iloc[:,1])

    for i in Providers_current_month_fin:
        if i in Providers_template_ncr or i == "Report Totals":
            continue
        else:
            new_providers.append(i)

    row_no, col_no = coordinate_finder_insheet_df(template_ar_worklist,"Providers List",1)

    template_ar_worklist.columns = template_ar_worklist.iloc[1,:]
    template_ar_worklist = template_ar_worklist.iloc[2:,:]

    current_month_providers = list(AR_report_df["Provider"].unique())
    template_providers = list(template_ar_worklist["Provider"].unique())

    new_providers_ar = []

    for i in current_month_providers:
        if i in template_providers:
            continue
        else:
            new_providers_ar.append(i)

#     new_providers_ar = ["Shanon","Ruby"]
    all_providers_ar = template_providers + new_providers_ar
    all_providers_ar = pd.DataFrame(all_providers_ar)

    Template = copy_paste_raw(Template,"AR WorkList",all_providers_ar,row_no+2,col_no+1,col_no+1)
    Template = left_right_border(Template,"AR WorkList",row_no+1,row_no+1+len(all_providers_ar),col_no+1)
    Template = left_right_bottom_border(Template,"AR WorkList",row_no+1+len(all_providers_ar),col_no+1,col_no+1)

    Template.save(file_saving_location)

    workbook = Workbook()
    workbook.LoadFromFile(file_saving_location)

    if len(new_providers_ar) > 0 :
        row_no, col_no = coordinate_finder_insheet_df(template_ar,"Parameter",2)
        insert_row_ar(workbook,template_ar,"AR",9,template_providers,new_providers_ar)
        insert_row_ar(workbook,template_ins_ar,"Insurance AR",3,template_providers,new_providers_ar)
        insert_row_ar(workbook,template_pat_ar,"Patient AR",3,template_providers,new_providers_ar)  

    row_no, col_no = coordinate_finder_insheet_df(template_ncr,"Provider",2)

    if len(new_providers) > 0 :
        Providers_template_ncr.remove('Overall')
        for i in new_providers:
            Providers_template_ncr.append(i)

        worksheet = workbook.Worksheets["NCR"]
        worksheet.InsertRow(row_no + len(Providers_template_ncr) - 3, len(new_providers),InsertOptionsType.FormatAsBefore)

    workbook.SaveToFile(file_saving_location)
    workbook.Dispose()

    app = xw.App(visible=False)
    wb = xw.Book(file_saving_location)
    wb.sheets['Evaluation Warning'].delete()
    wb.save(file_saving_location)
    wb.close()
    
    Providers = pd.DataFrame(Providers_template_ncr)
    
    Template=load_workbook(filename=file_saving_location)
    Template = copy_paste_raw(Template,"NCR",Providers,row_no+2,col_no+1,col_no+1)
    Template.save(file_saving_location)

    if len(new_providers_ar) > 0 :
        formatting_ar_sheets(template_providers,new_providers_ar,all_providers_ar,file_saving_location,"AR",template_ar,9)
        formatting_ar_sheets(template_providers,new_providers_ar,all_providers_ar,file_saving_location,"Insurance AR",template_ins_ar,3)
        formatting_ar_sheets(template_providers,new_providers_ar,all_providers_ar,file_saving_location,"Patient AR",template_pat_ar,3)

    print("Updated Output tabs")
    print("\n")
    
    
import time
start_time = time.time()

# Get the current system's username
username = getpass.getuser()
print("Username:", username)

#New user need to store this
path_folder = r"C:\Users\pragna_kandagatla\Desktop\Automation_Summary AR"

#Checking the files required i.e setup files required and storing
for filename in os.listdir(path_folder):
    file_path4 = os.path.join(path_folder, filename)
    if filename.endswith("Eye Associates of Tucson.xlsx"):
        template_path_eaot = file_path4
    elif filename.endswith("Tucson Ambulatory Anesthesia LLC.xlsx"):
        template_path_taa = file_path4
    elif filename.endswith("Wyatt Surgery Center.xlsx"):
        template_path_wsc = file_path4

folder_path=input("Paste the path of raw data folder= ")
current_month = pd.to_datetime(input("Please enter the start date of the month in (mm/dd/yyyy) format = "))
print("\n")
folder_path=folder_path.replace("\\","/").replace("\"","")

folder_name=folder_path.split("/")[-1]
file_saving_location_eaot = f"C:/Users/{username}/Downloads/Summary AR Report - Eye Associates of Tucson.xlsx"
file_saving_location_taa = f"C:/Users/{username}/Downloads/Summary AR Report - Tucson Ambulatory Anesthesia LLC.xlsx"
file_saving_location_wsc = f"C:/Users/{username}/Downloads/Summary AR Report - Wyatt Surgery Center.xlsx"

files_read_check = 1
for j, filename2 in enumerate(os.listdir(folder_path)):
    file_path2 = os.path.join(folder_path, filename2)
    for i, filename in enumerate(os.listdir(file_path2)):
        file_path = os.path.join(file_path2, filename)
        file_path=file_path.replace("\\","/").replace("\"","")
        read_df = read_raw_data(file_path)
        row_no_ar, col_ar = coordinate_finder_insheet_df(read_df, "Patient", 1)
        if row_no_ar > 11:
            if read_df.iloc[row_no_ar-11, col_ar] == "Practices: Wyatt Surgery Center":
                wsc_path = file_path2
            elif read_df.iloc[row_no_ar-11, col_ar] == "Practices: Eye Associates of Tucson":
                eaot_path = file_path2
            elif read_df.iloc[row_no_ar-11, col_ar] == "Practices: Tucson Ambulatory Anesthesia LLC":
                taa_path = file_path2
            
# After the loop, you can check the files assigned
print("Files assigned:")
print(f"wsc path: {wsc_path}")
print(f"eaot path: {eaot_path}")
print(f"taa path: {taa_path}")
print("\n")

files_read_check_wsc, wsc_ncr_path, wsc_ar_ins_path, wsc_ar_pat_path, wsc_fin_path = check_raw_data(wsc_path)
files_read_check_eaot, eaot_ncr_path, eaot_ar_ins_path, eaot_ar_pat_path, eaot_fin_path = check_raw_data(eaot_path)
files_read_check_taa, taa_ncr_path, taa_ar_ins_path, taa_ar_pat_path, taa_fin_path = check_raw_data(taa_path)

if not(files_read_check_wsc + files_read_check_eaot + files_read_check_taa) == 15 :
    print("Some of the Raw Data files are missing")
    sys.exit()

print("Executing WSC")
execute(template_path_wsc,file_saving_location_wsc,wsc_fin_path,wsc_ncr_path,wsc_ar_ins_path,wsc_ar_pat_path,current_month)
print("Executing TAA")
execute(template_path_taa,file_saving_location_taa,taa_fin_path,taa_ncr_path,taa_ar_ins_path,taa_ar_pat_path,current_month)
print("Executing EAOT")
execute(template_path_eaot,file_saving_location_eaot,eaot_fin_path,eaot_ncr_path,eaot_ar_ins_path,eaot_ar_pat_path,current_month)
print("Execution Done")


#Timer
end_time = time.time()
elapsed_time_seconds = end_time - start_time
elapsed_minutes = int(elapsed_time_seconds // 60)
elapsed_seconds = int(elapsed_time_seconds % 60)
print("Elapsed time: ", elapsed_minutes, "minutes", elapsed_seconds, "seconds")

