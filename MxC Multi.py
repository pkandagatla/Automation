#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import xlwings as xw
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.cell import get_column_letter
import sys
import getpass
from spire.xls.common import *
from spire.xls import *

def read_rawdata(filepath,cnames):
    
    #Gets the list of sheetnames
    sheet_names = pd.ExcelFile(filepath).sheet_names
    if(len(cnames)>0):
        #Creating a empty dataframe
        appended_df = pd.DataFrame(columns=cnames)

        # Iterate through each sheet and append data row by row
        for sheet_name in sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, names=cnames)  # Read data from each sheet with predefined column names
            appended_df = pd.concat([appended_df, df], ignore_index=True)  # Append rows to DataFrame

    else:
        appended_df = pd.DataFrame()

        # Iterate through each sheet and append data row by row
        for sheet_name in sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet_name)  # Read data from each sheet with predefined column names
            appended_df = pd.concat([appended_df, df], ignore_index=True)  # Append rows to DataFrame

    return(appended_df)

def concatenate_two_df(df1,df2):
    df2 = pd.concat([df1, df2], ignore_index=True)
    return df2
#Get the Facility name
def facility_name(df,x,y):
    facility_name=df.iloc[x,y]
    return facility_name

def reorder_column(df,columnname,position):
    df.insert(position,columnname,df.pop(columnname))
    return df

#It slices the dataframe if it finds a value from "x" space of the dataframe
def iter_row_slicer(df,valuetofind,x):
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                df=df.iloc[i+x:].reset_index(drop=True)
                found=True
                break
        if found is True:
            break
    if(found is False):
        message="No Value"
        return message
    return df

#New column creation if finds a value in the new column the value to be assigned before and after the values row indexes
def new_column_value(df,columnname,valuetofind,assign_before,assign_after):
    df[columnname]=0
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if df.iloc[i, j] == valuetofind:
                df.loc[:i, columnname] = assign_before
                df.loc[i+1:, columnname] = assign_after
                found=True
                break
        if found is True:
            break
    return df

#This function splits a datframe column into two by any character we say to split
def split_a_column(df,split_col,first_col_name,second_col_name,splitby):
    df[[first_col_name,second_col_name]]=df[split_col].str.split(splitby,expand = True)
    return df

def first_rowofdf(df):
    first_row=pd.DataFrame()
    df=pd.DataFrame(df)
    first_row = df.iloc[0:1, :] 
    return first_row

def assign_value_to_df(df,x,y,value):
    df.iloc[x,y]=value
    return df

#function to delete values that exist after the particular value
def delete_after_slicer(df,valuetofind,x):   
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                df=df.iloc[0:i+x,:].reset_index(drop=True)  
                found=True
                break
        if found is True:
            break
    return df

#Here we find a particular values coordinate in a sheet with the recurrence at what time
def coordinate_finder_insheet(df,valuetofind,recurrence):   
    found=False
    count=0
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                row_no=i+1
                col_no=j+1
                count+=1
                if(count==recurrence):
                    found=True
                    break
        if found is True:
            break 
    return(row_no,col_no)

#Here we find a particular values coordinate in a sheet with the recurrence at what time
def coordinate_finder_insheet_df(df, valuetofind, recurrence):   
    found = False
    count = 0
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if df.iloc[i, j] == valuetofind:
                row_no = i
                col_no = j
                count += 1
                if count == recurrence:
                    found = True
                    break
        if found:
            break
    if found:
        return (row_no, col_no)
    else:
        return (0, 0) 

#Here we are pasting the values from a dataframe to the raw data sheet
def copy_paste_raw(Template,sheetname,df,from_row,from_col,to_col):
    
    Sheet=Template[sheetname]

    row_number=from_row
    i=0
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row-1+len(df),min_col=from_col,max_col=to_col):
        j=0
        col_number=from_col
        for cell in row:
            Sheet.cell(row=row_number,column=col_number).value=df.iloc[i,j]
            j+=1
            col_number+=1
        i+=1
        row_number+=1 
    return(Template)

#Returns a dataframe converted from a excel sheet
def create_df_from_sheet(sheet_name,file_path):
    df=pd.read_excel(file_path,sheet_name=sheet_name)
    return df

#Find a value in the list with unique
def unique_values_except(df, omit_list, column_index):
    wanted_list = []
    for i in range(len(df)):
        value = df.iloc[i, column_index]
        if value not in omit_list and value not in wanted_list:
            wanted_list.append(value)
    return wanted_list

#Here we are extending the formulas till the end we have the values
def extend_formulas(Template,sheetname,df,from_row,from_col,to_col):
    Sheet=Template[sheetname]
    i=from_row
    for row in Sheet.iter_rows(min_row=from_row,max_row=from_row-2+len(df),min_col=from_col,max_col=to_col):
        j=from_col
        for cell in row:
            Sheet.cell(row=i,column=j).value=Translator(Sheet.cell(row=from_row-1,column=j).value,origin=Sheet.cell(row=from_row-1,column=j).coordinate).translate_formula(Sheet.cell(row=i,column=j).coordinate)    
            j+=1
        i+=1
    return(Template)

#Returns a dataframe converted from a excel sheet
def create_df_from_sheet(sheet_name,file_path):
    df=pd.read_excel(file_path,sheet_name=sheet_name)
    return df

#To get a dataframe where we need particular value of a column and it should satisfy the value from another column 
def get_row_with_condition(df,search_col,value_col,search_value):
    list1=[]
    for i in range(len(df)):
        if(df.iloc[i,search_col]==search_value):
            list1.append(df.iloc[i,value_col])
    df=pd.DataFrame(list1)
    return df

def concatenate_two_df(df1,df2):
    df2 = pd.concat([df1, df2], ignore_index=True)
    return df2

#To get a unique value of the dataframe by giving the column indices
def unique_row_values(df,col_no):
    df=df.iloc[:,col_no].unique()
    df=pd.DataFrame(df)
    return df

#It slices the dataframe if it finds a value from "x" space of the dataframe
def iter_row_slicer(df,valuetofind,x):
    found=False
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==valuetofind):
                df=df.iloc[i+x:].reset_index(drop=True)
                found=True
                break
        if found is True:
            break
    if(found is False):
        message="No Value"
        return message
    return df

#Triggers calculation in the sheet
# def perform_calc(filepath):
#     # Open the Excel workbook
#     wb = xw.Book(filepath)
#     #Make app not visible
#     app = xw.App(visible=False)
#     # Trigger calculation
#     wb.app.calculate()
#     # Save the workbook
#     wb.save()
#     # Close the workbook
#     wb.close()
    

def perform_calc(filepath):
    # Create an instance of the App with visibility off
    app = xw.App(visible=False)
    try:
        # Open the Excel workbook
        wb = app.books.open(filepath)
        # Trigger calculation
        wb.app.calculate()
        # Save the workbook
        wb.save()
        # Close the workbook
        wb.close()
    finally:
        # Quit the app to release resources
        app.quit()

    
#To trim a column 
def trim_column(df,column_no):
    df.iloc[:,column_no] = df.iloc[:,column_no].str.strip()
    return df

#To delete the specific cells in a range
def delete_cell_range(worksheet,start_row,start_col,end_row,end_col):
    if(start_col<end_col ):
        range_to_delete = worksheet.Range[start_row,start_col,end_row,end_col] 
        worksheet.DeleteRange(range_to_delete, DeleteOption.MoveLeft)
        return worksheet
        
#Deletes the below rows in the sheet 
def deleting_below_rows(worksheet,start_row,no_of_rows):
    worksheet.DeleteRow(start_row,no_of_rows)
    return worksheet

#Deletes the below rows in the sheet 
def deleting_right_columns(worksheet,start_col,no_of_cols):
    worksheet.DeleteColumn(start_col,no_of_cols)
    return worksheet

#Function to find the number of months between two dates
def to_find_no_months(start_date,end_date):
    dif=relativedelta(end_date,start_date)
    no_months= dif.years*12 + dif.months+1
    return no_months

#split a value and get as a list dataframe , value to find, the thing which used to split, row coordinate from the value we find and col as well 
def split_value(df,to_value,delimiter,x,y):
    for i in range(len(df)):
        stop=False
        for j in range(len(df.columns)):
            if(df.iloc[i,j]==to_value):
                date_range=df.iloc[i+x,j+y].split(delimiter)
                stop=True
                break
            else:
                continue
        if(stop is True):
            break
    return date_range

#Get the start and end range of the assessment
def assessment_range(date_range):
    start_date=date_range[0]
    start_date = datetime.strptime(start_date, '%m/%d/%Y')
    if(start_date.day!=1):
        start_date += relativedelta(months=1)
        start_date = start_date.replace(day=1)

    end_date=date_range[1]
    end_date = datetime.strptime(end_date, '%m/%d/%Y')

    if((end_date+timedelta(days=1)).day!=1):
        end_date -= relativedelta(months=1)
        end_date = end_date.replace(day=1)
        
    else:
        end_date = end_date.replace(day=1)

    return(start_date,end_date)

#Funtion to paste values
def paste_value_in_cell(workbook,sheet,row_no,col_no,paste_value):
    worksheet=workbook[sheet]
    worksheet.cell(row=row_no,column=col_no).value=paste_value
    return workbook

#Will add a bottom border for a row in a column range
def bottom_border_one_cell(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def right_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_right_bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def bottom_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style='medium'))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template

def left_right_border(Template,sheetname,row_no,col_start,col_end):
    sheet=Template[sheetname]
    # Create a border for the range of cells A1:C5
    border = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))

    for col in range(col_start, col_end+1):
        cell = sheet.cell(row=row_no, column=col)
        cell.border = border
        
    return Template
def get_cell_range(start_row, start_col, end_row, end_col):
    # Convert column indices to letters
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(end_col)
    
    # Construct the cell range string
    cell_range = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
    return cell_range

def cut_cell_range(Template,sheet_name,cell_range,row_to_move,col_to_move):
    ws=Template[sheet_name]
    ws.move_range(cell_range, rows=row_to_move, cols=col_to_move)
    return Template

def cut_cell_range_translate(Template,sheet_name,cell_range,row_to_move,col_to_move):
    ws=Template[sheet_name]
    ws.move_range(cell_range, rows=row_to_move, cols=col_to_move,translate=True)
    return Template

#Triggers calculation in the sheet
def perform_calc_delete_extra_sheet(filepath,sheetname):
    # Open the Excel workbook
    wb = xw.Book(filepath)
    #Make app not visible
    app = xw.App(visible=False)
    # Trigger calculation and deleting extra sheet
    wb.sheets[sheetname].delete()
    
    wb.app.calculate()
    wb.save()

    # Close the workbook
    wb.close()
    app.kill()

#This will apply a U border for a boundary we define
def border_apply_bottom_u_type(Template,sheetname,row_start,row_end,col_start,col_end):
    sheet=Template[sheetname]

    left_border = Border(left=Side(style='medium'), 
                right=Side(style=None), 
                top=Side(style=None), 
                bottom=Side(style=None))

    right_border = Border(left=Side(style=None), 
                right=Side(style="medium"), 
                top=Side(style=None), 
                bottom=Side(style=None))

    bottom_border = Border(left=Side(style=None), 
                right=Side(style=None), 
                top=Side(style=None), 
                bottom=Side(style="medium"))

    left_bottom_border = Border(left=Side(style='medium'), 
                     right=Side(style=None), 
                     top=Side(style=None), 
                     bottom=Side(style='medium'))

    right_bottom_border = Border(left=Side(style=None), 
                     right=Side(style="medium"), 
                     top=Side(style=None), 
                     bottom=Side(style='medium'))
    
    left_right_border = Border(left=Side(style="medium"), 
                     right=Side(style="medium"), 
                     top=Side(style=None), 
                     bottom=Side(style=None))
    
    left_right_bottom_border = Border(left=Side(style='medium'), 
                right=Side(style='medium'), 
                top=Side(style=None), 
                bottom=Side(style='medium'))

    if(col_start!=col_end):
        for i in range(row_start,row_end+1):
            cell1=sheet.cell(row=i,column=col_start)
            cell2=sheet.cell(row=i,column=col_end)
            cell1.border=left_border
            cell2.border=right_border
        for j in range(col_start+1,col_end):
            cell1=sheet.cell(row=row_end,column=j)
            cell1.border=bottom_border
        cell3=sheet.cell(row=row_end,column=col_start)
        cell3.border=left_bottom_border
        cell4=sheet.cell(row=row_end,column=col_end)
        cell4.border=right_bottom_border
        
    else:
        for i in range(row_start,row_end+1):
            cell1=sheet.cell(row=i,column=col_start)
            cell1.border=left_right_border
            
        cell3=sheet.cell(row=row_end,column=col_start)
        cell3.border=left_right_bottom_border
    
    return Template

T_pathx=r"C:\Users\pkandagatla\Desktop\br"
for filenamex in os.listdir(T_pathx):
    folder_path = os.path.join(T_pathx, filenamex)
    
    import time
    start_time = time.time()
    
    # Get the current system's username
    username = getpass.getuser()
    print("Username:", username)
    
    #Input of the file
    # folder_path = sys.argv[1]
    folder_path=folder_path.replace("\\","/").replace("\"","")
    
    #New user need to store this
    path_folder = r"C:\Users\pkandagatla\Downloads\Assessment Multi Automation\Assessment Multi Automation"
    
    #Checking the files required i.e setup files required and storing
    for filename in os.listdir(path_folder):
        file_path4 = os.path.join(path_folder, filename)
        if filename.endswith("AR Assessment - Multi Facility Template (MxC).xlsx"):
            template_path = file_path4
        elif filename.endswith("AR Assessment - Payer Grouping.xlsx"):
            payer_grp_send = file_path4
        elif filename.endswith("censor-beep-88052.mp3"):
            audio_file_path = file_path4
            
    # if __name__ == "__main__":
    #     folder_path=input("Paste the path of raw data folder= ")
    #     folder_path=folder_path.replace("\\","/").replace("\"","")
    
    folder_name=folder_path.split("/")[-1]
    file_saving_location = f"C:/Users/{username}/Downloads/AR Assessment - {folder_name}.xlsx"
    payer_grp_mail = f"C:/Users/{username}/Downloads/AR Assessment - Payer Grouping - {folder_name}.xlsx"
    
    files_read_check = 1
    for j, filename2 in enumerate(os.listdir(folder_path)):
        file_path2 = os.path.join(folder_path, filename2)
        if os.path.isfile(file_path2):
            read_df = pd.read_excel(file_path2, header=None, engine="openpyxl")
            row_no_payer, col_payer = coordinate_finder_insheet_df(read_df, "Payer", 1)
            if read_df.iloc[row_no_payer, col_payer + 1] == "Payer Group":
                payer_group = file_path2
                files_read_check += 1
        else:
                
            for i, filename in enumerate(os.listdir(file_path2)):
                if i > 0: break
                file_path = os.path.join(file_path2, filename)
                read_df = pd.read_excel(file_path, header=None)
                row_no_trans, col_trans = coordinate_finder_insheet_df(read_df, "Charges", 1)
                row_no_depos, col_depos = coordinate_finder_insheet_df(read_df, "Deposit Distribution Report:", 1)
                row_no_aging, col_aging = coordinate_finder_insheet_df(read_df, "Balance", 1)
                row_no_dso, col_dso = coordinate_finder_insheet_df(read_df, "Days of Sales Outstanding", 1)
    
                locals()[f"excel{j}"]  =read_df
                if read_df.iloc[row_no_trans + 1, col_trans] == "Transaction":
                    T_path = file_path2
                    files_read_check += 1
                elif read_df.iloc[row_no_depos + 2, col_depos] == "Payer":
                    D_path = file_path2
                    files_read_check += 1 
                elif read_df.iloc[row_no_dso + 2, col_dso] == "Facility ":
                    dso_path = file_path2
                    files_read_check += 1
                elif read_df.iloc[row_no_aging - 1, col_aging ] == "Outstanding":
                    A_path = file_path2
                    files_read_check += 1
            
    
    if not(files_read_check>5):
        print(f"Some of the raw data files are missing in this folder = {folder_path}")
        sys.exit()
    # After the loop, you can check the files assigned
    print("Files assigned:")
    print(f"Transaction Report: {T_path}")
    print(f"Deposit Report: {D_path}")
    print(f"DSO Report: {dso_path}")
    print(f"Aging report: {A_path}")
    print(f"Payer group: {payer_group}")
    
    #storing all the sheets as a dataframe
    template_transaction=pd.read_excel(template_path,sheet_name="Transaction",header=None)
    template_deposit=pd.read_excel(template_path,sheet_name="Deposit",header=None)
    template_aging=pd.read_excel(template_path,sheet_name="Agings",header=None)
    template_dso=pd.read_excel(template_path,sheet_name="DSO",header=None)
    template_grouping=pd.read_excel(template_path,sheet_name="Grouping",header=None)
    template_payments=pd.read_excel(template_path,sheet_name="Payments",header=None)
    template_ncr_cal=pd.read_excel(template_path,sheet_name="Payer-Wise NCR Cal",header=None)
    template_aging_cal=pd.read_excel(template_path,sheet_name="Aging Cal",header=None)
    template_clean_up=pd.read_excel(template_path,sheet_name="Clean Up estimates",header=None)
    template_dsoncr=pd.read_excel(template_path,sheet_name="DSO & NCR",header=None)
    template_assessment_metrics=pd.read_excel(template_path,sheet_name="Assessment Metrics",header=None)
    template_ncr_facility=pd.read_excel(template_path,sheet_name="NCR By Facility",header=None)
    template_ncr_payer=pd.read_excel(template_path,sheet_name="NCR By Payer",header=None)
    
    print("Template files are converted to dataframe")
    #Loading the template
    Template=load_workbook(filename=template_path)
    #-----------------------------------------Dataframe Creation for Raw Data--------------------------------------------
    #----------- Transaction Report
    transaction_df=pd.DataFrame()
    facility_names=[]
    transaction_date_range=[]
    transaction_facility_name=[]
    for filename in os.listdir(T_path):
        file_path = os.path.join(T_path, filename)
    
        TreportCnames = ["Transaction Date", "Description", "Status", "Service From Date", "Service Through Date", "Units", "Empty", "Days", "Amounts"]
    
        #Passing Transaction report file path to retrieve dataframe of it
        transaction_report_df= read_rawdata(file_path,TreportCnames)
        transaction_report_df2=transaction_report_df
        
        #Getting the daterange of the assessment
        t_date_range=transaction_report_df2.iloc[2,0]
        start_index = t_date_range.find(":") + 2
        end_index = t_date_range.find("    ")
        t_date_range = t_date_range[start_index:end_index]
        transaction_date_range.append(t_date_range)
    
        #Get facility name
        T_Facility_Name=facility_name(transaction_report_df,0,0)
        transaction_report_df["Facility Name"]=T_Facility_Name
        facility_names.append(T_Facility_Name)
        transaction_facility_name.append(T_Facility_Name)
    
        #reorder column
        transaction_report_df=reorder_column(transaction_report_df,"Facility Name",0)
    
        #Row slice if it finds the Date
        transaction_report_df=iter_row_slicer(transaction_report_df,"Date",1)
    
        #Test column creation
        transaction_report_df=new_column_value(transaction_report_df,"Test","Date",1,0)
    
        #Concatenate all dataframe
        transaction_df=concatenate_two_df(transaction_df,transaction_report_df)
    print("Read Transaction report")
    print(len(transaction_df))
    #-------------------------Deposit Report--------------------
    deposit_df=pd.DataFrame()
    deposit_date_range=[]
    deposit_facility_name=[]
    finder_in=0
    for filename in os.listdir(D_path):
            file_path = os.path.join(D_path, filename)
            
            DreportCnames = ["Payer","Emply 1" ,"Empty 2","MR#", "Charge Date", "Cash Receipe date", "From Through Date", "Distributed Amount", "Adjustments"]
        
            #Passing Deposit report file path to retrieve dataframe of it 
            Deposit_report_df=read_rawdata(file_path,DreportCnames)
            
            #Retreving date range for all the facilities
            Deposit_report_df3=Deposit_report_df
            d_date_range= Deposit_report_df3.iloc[4,3]
            deposit_date_range.append(d_date_range)
            
            
            if(finder_in==0):
                Deposit_report_df_2=Deposit_report_df
            finder_in+=1
    
            #Get facility name
            D_Facility_Name=facility_name(Deposit_report_df,2,0)
            deposit_facility_name.append(D_Facility_Name)
            Deposit_report_df["Facility Name"]=D_Facility_Name
    
            #reorder column
            Deposit_report_df=reorder_column(Deposit_report_df,"Facility Name",0)
    
            #Row slice if it finds the Deposit distribution
            Deposit_report_df=iter_row_slicer(Deposit_report_df,"Deposit Distribution Report:",5)
    
            #Split Column
            Deposit_report_df=split_a_column(Deposit_report_df,"From Through Date","From Date","Through Date"," - ")
            Deposit_report_df['From Date'] = pd.to_datetime(Deposit_report_df['From Date'])
            Deposit_report_df['Through Date'] = pd.to_datetime(Deposit_report_df['Through Date'])
    
            
            #Concatenate all dataframe
            deposit_df=concatenate_two_df(deposit_df,Deposit_report_df)
    
    print("Read Deposit report")
    
    aging_df=pd.DataFrame()
    aging_dict={}
    aging_dict2={}
    aging_raw_file_count=0
    aging_payers=pd.DataFrame()
    aging_payers_list = pd.DataFrame()
    first_file=0
    aging_facility_names=[]
    for filename in os.listdir(A_path):
            file_path = os.path.join(A_path, filename)
            
            #No predefined column names so emptytt list
            AreportCnames=[]
    
            #Passing Aging report file path to retrieve dataframe of it
            Aging_report_df=read_rawdata(file_path,AreportCnames)
    
            #Get facility name
            A_Facility_Name=facility_name(Aging_report_df,0,0)
            aging_facility_names.append(A_Facility_Name)
            Aging_report_df["Facility Name"]=A_Facility_Name
            aging_dict2[f"aging_raw_data{aging_raw_file_count}"]=Aging_report_df
    
            #reorder column
            Aging_report_df=reorder_column(Aging_report_df,"Facility Name",0)
    
            #Slice for Resident aging here we need first column to know the aging buckets
            Aging_report_df=iter_row_slicer(Aging_report_df,"Balance",0)
    
            #First row
            Age_bucket=pd.DataFrame()
            Age_bucket=first_rowofdf(Aging_report_df)
    
            #Slice for Payer Type Summary 
            Aging_report_df=iter_row_slicer(Aging_report_df,"Payer Type Summary",1)
    
            #Concatnate two DFs
            Aging_report_df=concatenate_two_df(Age_bucket,Aging_report_df)
    
            #Give values to some postion in dfs
            Aging_report_df=assign_value_to_df(Aging_report_df,0,0,"Facility Name")
            Aging_report_df=assign_value_to_df(Aging_report_df,0,1,"Payer Type")
    
            #function to delete values that exist after the particular value
            Aging_report_df=delete_after_slicer(Aging_report_df,"Payer Type Total",1)
    
            #Drop third row
            Aging_report_df.drop(Aging_report_df.columns[2], axis=1, inplace=True)
            
            #Aging payers unique
            aging_payers=Aging_report_df.iloc[1:len(Aging_report_df)-1,]
            aging_payers=unique_row_values(aging_payers,1)
            aging_payers_list=concatenate_two_df(aging_payers,aging_payers_list)
            aging_payers_list=unique_row_values(aging_payers_list,0)
    
            # Get the count of unique values
            no_aging_payers = len(aging_payers_list)
            
            aging_dict[f"aging_raw_data{aging_raw_file_count}"]=Aging_report_df
            aging_raw_file_count+=1
            
            #the first file alone the header will be there others we dont need
            if(first_file>0):
                Aging_report_df=Aging_report_df.iloc[1:,]
        
            first_file+=1
            #Concatenate with master dataframe
            aging_df=concatenate_two_df(aging_df,Aging_report_df)
            
    print("Read Aging reports")     
    
    #----------------------------DSO----------------------------------
            
    dso_df=pd.DataFrame()
    dso_facility_names=[]
    first_file=0
    dso_dict={}
    for filename in os.listdir(dso_path):
            file_path = os.path.join(dso_path, filename)
            
            #No predefined column names so emptytt list
            DSOreportCnames=[]
    
            #Passing DSO report file path to retrieve dataframe of it
            DSO_report_df=read_rawdata(file_path,DSOreportCnames)
    
            #Get facility name
            DSO_Facility_Name=facility_name(DSO_report_df,1,0)
            DSO_report_df["Facility Name"]=DSO_Facility_Name 
    
            #reorder column
            DSO_report_df=reorder_column(DSO_report_df,"Facility Name",0)
    
            #Slice for Days Sales of Outstanding
            DSO_report_df=iter_row_slicer(DSO_report_df,"Days of Sales Outstanding",2)
    
            # Slice for before
            DSO_report_df=delete_after_slicer(DSO_report_df,"MatrixCare Report",-1)
            
            DSO_report_df=DSO_report_df.iloc[:,1:]
            
            #Get facility Name
            dso_facility_names.append(DSO_report_df.iloc[1,0])
          
            
            dso_dict[f"dso_raw_data{first_file}"]=DSO_report_df
            
            #the first file alone the header will be there others we dont need
            if(first_file>0):
                DSO_report_df=DSO_report_df.iloc[1:,]  
                
            first_file+=1
            
            dso_df=concatenate_two_df(dso_df,DSO_report_df)
            
    print("Read DSO report") 
    
    #--------------------------------------Raw Data Check-------------------------------
    #-----------------RAW DATA CHECKS---------------------------------
    # Initial Raw Data checks
    #Facility check----------------------------
    Facility_access = "Yes"
    
    if set(dso_facility_names) == set(aging_facility_names) == set(deposit_facility_name) == set(transaction_facility_name):
        pass
    else:
        print("Transaction Report= ", transaction_facility_name)
        print("Deposit Report= ", deposit_facility_name)
        print("Aging Report= ", aging_facility_names)
        print("DSO Report= ", dso_facility_names)
        Facility_access = input(r"Facility Names are mismatching between reports, is it fine to proceed? (Yes/No)= ")
        print("We can dig more in which reports we have mismatch of Facility names")
        if Facility_access != "Yes":
            print("Please use corresponding files for Assessment")
    
    facility_name_check_transaction = 0
    
    if len(set(transaction_facility_name)) == first_file:
        pass
    else:
        facility_name_check_transaction += 1
        print("We have different facility names for files of transaction report")
        print(transaction_facility_name)
    
    facility_name_check_deposit = 0
    if len(set(deposit_facility_name)) == first_file:
        pass
    else:
        facility_name_check_deposit += 1
        print("We have different facility names for files of deposit report")
        print(deposit_facility_name)
    
    facility_name_check_aging = 0
    if len(set(aging_facility_names)) == first_file:
        pass
    else:
        facility_name_check_aging += 1
        print("We have different facility names for files of aging report")
        print(aging_facility_names)
    
    facility_name_check_dso = 0
    if len(set(dso_facility_names)) == first_file:
        pass
    else:
        facility_name_check_dso += 1
        print("We have different facility names for files of dso report")
        print(dso_facility_names)
    
    # Checking for Date range of deposit and transaction report
    Date_range_access = "Yes"
    transaction_deposit_date_error = 0
    if set(deposit_date_range) != set(transaction_date_range):
        print("Transaction and deposit report date range do not match")
        for i in range(len(transaction_date_range)):
            print(transaction_facility_name[i])
            print("Transaction = ", transaction_date_range[i])
            print("Deposit = ", deposit_date_range[i])
            transaction_deposit_date_error += 1
        Date_range_access = input(r"Do you wish to proceed with this date range (Yes/No)= ")
        print("We can dig more for which reports we have mismatch in facility names")
        if Date_range_access != "Yes":
            print("Get the correct file either Transaction or Deposit to run the assessment")
    
    transaction_date_error = 0
    if len(set(transaction_date_range)) != 1:
        transaction_date_error = 1
    if transaction_date_error > 0:
        print("Date range in one or more of the facilities are different in transaction report, please refer below")
        for i in range(len(transaction_date_range)):
            print(transaction_facility_name[i], " has date range ", transaction_date_range[i])
    print("Transaction Date range= ",set(deposit_date_range))

    
    deposit_date_error = 0
    if len(set(deposit_date_range)) != 1:
        deposit_date_error = 1
    if deposit_date_error > 0:
        print("Date range in one or more of the facilities are different in deposit report, please refer below")
        for i in range(len(deposit_date_range)):
            print(deposit_facility_name[i], " has date range ", deposit_date_range[i])
    print("Deposit Date range= ",set(deposit_date_range))
            
    g=0
    dso_access="Yes"
    dso_mismatch=0
    for i in dso_dict:
        if(g>0):
            k=set(dso_dict[i].iloc[0])
            if(k!=j):
                dso_mismatch+=1
        j=set(dso_dict[i].iloc[0])
        g+=1
    if(dso_mismatch>0):
        print("The DSO months are not matching b/w facilities:")
        k=0
        for i in dso_dict:
            print("For facility",dso_facility_names[k])
            print(dso_dict[i].iloc[0])
            k+=1
        dso_access = input(r"Do you wish to proceed with this date range for DSO (Yes/No)= ")
        if dso_access != "Yes":
            print("Get the correct file either Transaction or Deposit to run the assessment")
    
    
    #Aging check---------------
    aging_headers_check=0
    aging_access="Yes"
    aging_access1="Yes"
    for aging_check in aging_dict:
        check=0
        Aging_report_dfheaders=["Facility Name","Payer Type","Balance","0-30","31-60","61-90","91-120","121-150","151-180","181-210","211-365",">365"]
        if(len(Aging_report_dfheaders)!=len(aging_dict[aging_check].columns)):
            aging_headers_check+=1
            print(f"For facility {aging_dict[aging_check].iloc[1,0]} we have the age buckets as=")
            print(aging_dict[aging_check].iloc[0,:])
        else:
            for i in range(len(Aging_report_dfheaders)):
                if(Aging_report_dfheaders[i]==aging_dict[aging_check].iloc[0,i]):
                    check+=1
            if(check!=11):
                aging_headers_check+=1
                print(f"For facility {aging_dict[aging_check].iloc[1,0]} we have different age buckets as=")
                print(aging_dict[aging_check].iloc[0,:])
    if(aging_headers_check>0):
        aging_access1=input(r"Can we proceed with this Age buckets as they are in unsual format, can we continue assessment(Yes/No)= ")
        if(aging_access1!="Yes"):
            print("Please use the correct Aging report for the Assessment")
    
    iterator=-1
    aging_dates_check=0
    aging_date_check_input="Yes"
    for aging_check in aging_dict2:
        #Age Through and Cash Through Date Check
        iterator+=1
        age_through_date=aging_dict2[aging_check].iloc[2,1]
        cash_through_date=aging_dict2[aging_check].iloc[3,1]
        age_through_date = age_through_date.split(":")[1].strip()
        cash_through_date = cash_through_date.split(":")[1].strip()
        print("Age Through Date = ",facility_names[iterator]," = ",age_through_date)
        print("Cash Through Date = ",facility_names[iterator]," = ",cash_through_date)
        if(age_through_date!=cash_through_date):
            print(f"Age Through Date and Cash Through Date are not matching for facility {facility_names[iterator]}")
            print("Age Thru Date= ",age_through_date,"\n Cash Through Date= ",cash_through_date)
            aging_dates_check+=1
    
    if(aging_dates_check>0):              
        aging_date_check_input=input(r"Do you want to continue with this age through and cash through date?(Yes/No)")
        if(aging_date_check_input!="Yes"):
            print("Get the corrected Aging report to run the assessment")  
            
    #Check for Payer Type Summary
    aging_payer_summary_check=0
    aging_access3="Yes"
    for aging_check in aging_dict2:
        aging_check2=iter_row_slicer(aging_dict2[aging_check],"Payer Type Summary",1)
        if(len(aging_check2)==16):
            print(f"We don't have Payer Type Summary for the Facility {payer_type_summary}")
            aging_payer_summary_check+=1
            
    if(aging_payer_summary_check>0):              
            print("Get the corrected Aging report to run the assessment")
            aging_access3="No"
    
    # #Assessment date ranges
    # asses_date_range=split_value(Deposit_report_df_2,"Deposit Distribution Report:"," - ",0,3)
    # start_date,end_date=assessment_range(asses_date_range)
    
    # end_date_eomonth = end_date + timedelta(days=1)
    #         end_date += relativedelta(months=1)
    #         end_date = end_date.replace(day=1)
    
    # if(age_through_date==end_date):
        
        
        
    
    if(aging_access1==aging_access3==aging_date_check_input=="Yes"):
        aging_access="Yes"
    else:
        aging_access="No"
        print("Continue the Assessment with correct raw data files for Aging")
            
        
    if(aging_access==Facility_access==Date_range_access==dso_access=="Yes"):
        print("Assessment Satisfied all the raw data checks!")
    else:
        print("Assessment is being quit, Please comeback with appropriate raw data files")
        sys.exit()
    
    # transaction_report2split=transaction_report_df
    # len_t_report=len(transaction_report_df)
    # for i in range(0,len_t_report,500000):
    #     if(len_t_report-i<500000):
    #         start_range=i
    #         end_range=len_t_report-1
    #     else:
    #         start_range=i
    #         end_range=i+500000
    #     print(start_range,end_range)
    #     transaction_report_df=transaction_report2split.iloc[start_range:end_range]
    #     file_saving_location = f"C:/Users/{username}/Downloads/AR Assessment - {folder_name,start_range}.xlsx"
    #     payer_grp_mail = f"C:/Users/{username}/Downloads/AR Assessment - {folder_name,start_range} - Payer Grouping.xlsx"
    
    #------------AR Clean Up Estimates--------
    Aging_report_df2=aging_df.iloc[1:len(aging_df)-1,:]
    
    #Unique list of payers need to be added in the AR Clean up estimates
    wanted_list=unique_values_except(Aging_report_df2, ["Private", "Other", "Miscellaneous","Payer Type Total"], 1)
    wanted_list=pd.DataFrame(wanted_list)
    
    no_of_clean_up_payers=len(wanted_list)
    print("Unique Clean Up estimates payers retrieved")
    #--------------Payer Grouping--------------------------
    #Payer Group Dataframe
    pnames=[]
    payer_group_df= read_rawdata(payer_group,pnames)
    payer_group_df.columns=payer_group_df.iloc[0]
    
    #Slice for Payer
    payer_group_df=iter_row_slicer(payer_group_df,"Payer",1)
    payer_group_df=payer_group_df.iloc[:,1:3]
    print("Unique payer groups retreived for NCR Calculation")
    
    #Getting value for number of payers in the assessment
    payer_group_df=trim_column(payer_group_df,1)
    uniques_payer_groups=unique_row_values(payer_group_df,1)
    uniques_payer_groups.columns=["Payer Group"]
    uniques_payer_groups= uniques_payer_groups[uniques_payer_groups['Payer Group'] != 'Non-QHCR Billed Payer']
    no_of_payer_grps=len(uniques_payer_groups)
    
    #-------------------DSO & NCR Date range-------------------------
    #Passing the value
    asses_date_range=split_value(Deposit_report_df_2,"Deposit Distribution Report:"," - ",0,3)
    
    #Assessment date ranges
    start_date,end_date=assessment_range(asses_date_range)
    
    #to get number of months
    no_months=to_find_no_months(start_date,end_date)
    
    print("Deletion of extra rows and columns in Template started.....")
    #--------------------------------Deletion of Extra rows for Calculation sheets Start---------------------
    #creating variables for Spire libraries worksheets

    
    workbook = Workbook()
    workbook.LoadFromFile(template_path)
    
    worksheet_aging_cal = workbook.Worksheets["Aging Cal"]   
    worksheet_payments= workbook.Worksheets["Payments"] 
    worksheet_clean_up= workbook.Worksheets["Clean Up estimates"] 
    worksheet_ncr_cal= workbook.Worksheets["Payer-Wise NCR Cal"] 
    worksheet_ncr_facility= workbook.Worksheets["NCR By Facility"] 
    worksheet_ncr_payer= workbook.Worksheets["NCR By Payer"] 
    worksheet_ar_assessment= workbook.Worksheets["Assessment Metrics"]
    worksheet_dsoncr= workbook.Worksheets["DSO & NCR"]
    worksheet_dso= workbook.Worksheets["DSO"]
    worksheet_aging= workbook.Worksheets["Agings"]
    
    #----------------Aging Cal Deleting extra payers--------------------------
    #Find the position of the particular excel paths iterating over each table and deleting the extra payers here
    row_reducing=0
    
    #first file is the number of facilities variable
    
    for i in range(1,first_file+1+1):
        row_no,col_no=coordinate_finder_insheet(template_aging_cal,"Payer Type Summary",i)
        end_row_no,end_col_no=coordinate_finder_insheet(template_aging_cal,"Payer Type Total",i)
        
        start_row=row_no+no_aging_payers+1-row_reducing
        no_of_rows_delete=end_row_no-start_row-row_reducing
        
        worksheet_aging_cal=deleting_below_rows(worksheet_aging_cal,start_row,no_of_rows_delete)
        row_reducing+=no_of_rows_delete
        
    #Deleting the Aging cal extra facilities
    row_no_fac,col_no_fac=coordinate_finder_insheet(template_aging_cal,"Payer Type Summary",first_file+2)
    
    start_row_fac=row_no_fac-row_reducing
    worksheet_aging_cal=deleting_below_rows(worksheet_aging_cal,start_row_fac,10000)
    
    print("Aging cal deletion completed")
    
    # ----------------------Deletion for Clean up estimates-------------------
    #Find the position of the particular excel paths iterating over each table and deleting the extra payers here
    row_reducing=0
    
    #first file is the number of facilities variable
    
    for i in range(2,first_file+1+1):
        row_no,col_no=coordinate_finder_insheet(template_clean_up,"Payer Type Summary",i)
        end_row_no,end_col_no=coordinate_finder_insheet(template_clean_up,"Payer Type Total",i)
        
        start_row=row_no+no_of_clean_up_payers+1-row_reducing
        no_of_rows_delete=end_row_no-start_row-row_reducing
        
        worksheet_clean_up=deleting_below_rows(worksheet_clean_up,start_row,no_of_rows_delete)
        row_reducing+=no_of_rows_delete
        
    #Deleting the Clean up Estimates extra facilities
    row_no_fac,col_no_fac=coordinate_finder_insheet(template_clean_up,"Payer Type Summary",first_file+2)
    
    start_row_fac=row_no_fac-row_reducing
    worksheet_clean_up=deleting_below_rows(worksheet_clean_up,start_row_fac,10000)
    
    #Deleting extra facilities in the Clean up estimates for Overall %Collectable
    row_no_fac,col_no_fac=coordinate_finder_insheet(template_clean_up,"Facility Wise Summary",1)
    row_no_end,col_no_end=coordinate_finder_insheet(template_clean_up,"Facility Wise Totals",1)
    
    start_row_fac=row_no_fac+first_file+1
    no_of_rows=row_no_end-start_row_fac
    worksheet_clean_up=deleting_below_rows(worksheet_clean_up,start_row_fac,no_of_rows)
    
    row_no,col_no=coordinate_finder_insheet(template_clean_up,"Payer Type Summary",1)
    end_row_no,end_col_no=coordinate_finder_insheet(template_clean_up,"Payer Type Total",1)
    
    start_row=row_no+no_of_clean_up_payers+1
    no_of_rows_delete=end_row_no-start_row
    
    worksheet_clean_up=deleting_below_rows(worksheet_clean_up,start_row,no_of_rows_delete)
    
    print("AR Clean up estimates deletion completed")
    
    # ----------------------Deletion for Payments Cal-------------------
    #Find the position of the particular excel paths iterating over each table and deleting the extra payers here
    row_reducing=0
    
    #first file is the number of facilities variable
    
    for i in range(1,first_file+1):
        row_no,col_no=coordinate_finder_insheet(template_payments,"Payer Group",i)
        end_row_no,end_col_no=coordinate_finder_insheet(template_payments,"Overall",i)
        
        start_row=row_no+no_of_payer_grps+1-row_reducing
        no_of_rows_delete=end_row_no-start_row-row_reducing
        
        worksheet_payments=deleting_below_rows(worksheet_payments,start_row,no_of_rows_delete)
        row_reducing+=no_of_rows_delete
        
    #Deleting the Aging cal extra facilities
    row_no_fac,col_no_fac=coordinate_finder_insheet(template_payments,"Payer Group",first_file+1)
    
    start_row_fac=row_no_fac-row_reducing
    worksheet_payments=deleting_below_rows(worksheet_payments,start_row_fac,10000)
    
    print("Payments Cal deletion completed")   
    #----------------NCR Facility Wise--------------------------
    #Deletion for the number of months valid in the assessment
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_ncr_facility,"Payer Type",1)
    
    # worksheet_ncr_facility=delete_cell_range(worksheet_ncr_facility,row_no,col_no+1,10000,col_no+18-no_months)
    if(no_months>18):
        no_of_cols_to_delete=0
    else:
        no_of_cols_to_delete=18-no_months
    worksheet_ncr_facility=deleting_right_columns(worksheet_ncr_facility,col_no+1,no_of_cols_to_delete)
    
    #Find the position of the particular excel paths iterating over each table and deleting the extra payers here
    row_reducing=0
    
    #first file is the number of facilities variable
    
    for i in range(1,first_file+1+1):
        row_no,col_no=coordinate_finder_insheet(template_ncr_facility,"Payer Type",i)
        end_row_no,end_col_no=coordinate_finder_insheet(template_ncr_facility,"Overall",i)
        
        start_row=row_no+no_of_payer_grps-row_reducing+1
        no_of_rows_delete=end_row_no-start_row-row_reducing
        
        worksheet_ncr_facility=deleting_below_rows(worksheet_ncr_facility,start_row,no_of_rows_delete)
        row_reducing+=no_of_rows_delete
        
    #Deleting the Aging cal extra facilities
    row_no_fac,col_no_fac=coordinate_finder_insheet(template_ncr_facility,"Payer Type",first_file+2)
    
    start_row_fac=row_no_fac-row_reducing
    worksheet_ncr_facility=deleting_below_rows(worksheet_ncr_facility,start_row_fac,10000)
    print("NCR Facility wise deletion completed")
        
    
    # ----------------------Deletion for NCR Payer-------------------
    
    #Deletion for the number of months valid in the assessment
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_ncr_payer,"Facility",1)
    
    # worksheet_ncr_payer=delete_cell_range(worksheet_ncr_payer,row_no,col_no+1,10000,col_no+18-no_months)
    if(no_months>18):
        no_of_cols_to_delete=0
    else:
        no_of_cols_to_delete=18-no_months
    worksheet_ncr_payer=deleting_right_columns(worksheet_ncr_payer,col_no+1,no_of_cols_to_delete)
    
    #Find the position of the particular excel paths iterating over each table and deleting the extra payers here
    row_reducing=0
    
    #first file is the number of facilities variable
    
    for i in range(1,no_of_payer_grps+1):
        row_no,col_no=coordinate_finder_insheet(template_ncr_payer,"Facility",i)
        end_row_no,end_col_no=coordinate_finder_insheet(template_ncr_payer,"Overall",i)
        
        start_row=row_no+first_file+1-row_reducing
        no_of_rows_delete=end_row_no-start_row-row_reducing
        
        worksheet_ncr_payer=deleting_below_rows(worksheet_ncr_payer,start_row,no_of_rows_delete)
        row_reducing+=no_of_rows_delete
        
    #Deleting the Aging cal extra facilities
    row_no_fac,col_no_fac=coordinate_finder_insheet(template_ncr_payer,"Facility",no_of_payer_grps+1)
    
    start_row_fac=row_no_fac-row_reducing
    worksheet_ncr_payer=deleting_below_rows(worksheet_ncr_payer,start_row_fac,10000)
    
    print("NCR Payer wise deletion completed")
    
    # ----------------------Deletion for NCR Cal-------------------
    #Deletion for the number of months valid in the assessment
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_ncr_cal,"Metrics",1)
    
    # worksheet_ncr_cal=delete_cell_range(worksheet_ncr_cal,row_no-2,col_no+1,10000,col_no+18-no_months)
    if(no_months>18):
        no_of_cols_to_delete=0
    else:
        no_of_cols_to_delete=18-no_months
    worksheet_ncr_cal=deleting_right_columns(worksheet_ncr_cal,col_no+1,no_of_cols_to_delete)
    
    
    #Find the position of the particular excel paths iterating over each table and deleting the extra payers here
    row_reducing=0
    
    #first file is the number of facilities variable
    row_no_initial,col_no_initial=coordinate_finder_insheet(template_ncr_cal,"Metrics",1)
    row_no_payer_initial,col_no_payer_initial=coordinate_finder_insheet(template_ncr_cal,"Expected Revenue - Overall",no_of_payer_grps+1)
    rows_to_keep=row_no_payer_initial-row_no_initial
    
    for i in range(1,first_file+1+1):
        row_no,col_no=coordinate_finder_insheet(template_ncr_cal,"Metrics",i)
        end_row,end_col=coordinate_finder_insheet(template_ncr_cal,"Metrics",i+1)
           
        start_row=row_no+rows_to_keep-row_reducing
        no_of_rows_delete=end_row-start_row-row_reducing-1
    
        worksheet_ncr_cal=deleting_below_rows(worksheet_ncr_cal,start_row,no_of_rows_delete)
        row_reducing+=no_of_rows_delete
    
    # #Deleting the extra facilities
    row_no_fac,col_no_fac=coordinate_finder_insheet(template_ncr_cal,"Metrics",first_file+1+1)
    start_row_fac=row_no_fac-row_reducing
    worksheet_ncr_cal=deleting_below_rows(worksheet_ncr_cal,start_row_fac,10000)
    
    print("Deletion of NCR Cal completed")
    
    #--------AR Assessment Metrics------------
    row_no,col_no=coordinate_finder_insheet(template_assessment_metrics,"Parameters",1)
    
    start_col=col_no+2+first_file
    worksheet_ar_assessment=deleting_right_columns(worksheet_ar_assessment,start_col,30)
    
    #--------DSO & NCR------------
    #Deletion of extra facilities
    row_no,col_no=coordinate_finder_insheet(template_dsoncr,"Metrics",first_file+2)
    
    start_row=row_no
    
    worksheet_dsoncr=deleting_below_rows(worksheet_dsoncr,start_row,100000)
    #Deletion for the number of months valid in the assessment
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_dsoncr,"Metrics",1)
    
    # worksheet_dsoncr=delete_cell_range(worksheet_dsoncr,row_no-2,col_no+2,10000,col_no+18-no_months)
    if(no_months>18):
        no_of_cols_to_delete=0
    else:
        no_of_cols_to_delete=18-no_months
    worksheet_dsoncr=deleting_right_columns(worksheet_dsoncr,col_no+1,no_of_cols_to_delete)
    
    #Deletion of extra facilities in DSO & NCR 
    row_no,col_no=coordinate_finder_insheet(template_dsoncr,"Facility Name",1)
    
    start_row=row_no+1+first_file
    no_of_rows=30-first_file
    worksheet_dsoncr=deleting_below_rows(worksheet_dsoncr,start_row,no_of_rows)
    
    #-------------DSO Raw data deletion-------------
    row_no,col_no=coordinate_finder_insheet(template_dso,"Days of Sales Outstanding",1)
    start_row=row_no+3+first_file
    
    worksheet_dso=deleting_below_rows(worksheet_dso,start_row,1000)
    
    #-------------Aging Raw data deletion-------------
    row_no,col_no=coordinate_finder_insheet(template_aging,"Facility Name",1)
    start_row=row_no+len(aging_df)
    
    worksheet_aging=deleting_below_rows(worksheet_aging,start_row,1000)
    
    print("Deletion of other output tabs completed")
    
    workbook.SaveToFile(file_saving_location)
    workbook.Dispose()
    print("Triggering Calculation")
    perform_calc(file_saving_location)
    
    border_transaction=pd.read_excel(file_saving_location,sheet_name="Transaction",header=None)
    border_deposit=pd.read_excel(file_saving_location,sheet_name="Deposit",header=None)
    border_aging=pd.read_excel(file_saving_location,sheet_name="Agings",header=None)
    border_dso=pd.read_excel(file_saving_location,sheet_name="DSO",header=None)
    border_grouping=pd.read_excel(file_saving_location,sheet_name="Grouping",header=None)
    border_payments=pd.read_excel(file_saving_location,sheet_name="Payments",header=None)
    border_ncr_cal=pd.read_excel(file_saving_location,sheet_name="Payer-Wise NCR Cal",header=None)
    border_aging_cal=pd.read_excel(file_saving_location,sheet_name="Aging Cal",header=None)
    border_clean_up=pd.read_excel(file_saving_location,sheet_name="Clean Up estimates",header=None)
    border_dsoncr=pd.read_excel(file_saving_location,sheet_name="DSO & NCR",header=None)
    border_assessment_metrics=pd.read_excel(file_saving_location,sheet_name="Assessment Metrics",header=None)
    border_ncr_facility=pd.read_excel(file_saving_location,sheet_name="NCR By Facility",header=None)
    border_ncr_payer=pd.read_excel(file_saving_location,sheet_name="NCR By Payer",header=None)
    
    #Loading the workbook again
    Template=load_workbook(filename=file_saving_location)
    print("Pasting of Raw data values begun.....")        
    #-------------------Pasting Values--------------------------------
    #---Transaction Report-----------
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_transaction,"Facility",1)
    to_row,to_col=coordinate_finder_insheet(template_transaction,"Test",1)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Transaction",transaction_df,row_no+1,col_no,to_col)
    
    #Formula pasting range finder
    for_row_no,for_col_no=coordinate_finder_insheet(template_transaction,"SOM Service through date",1)
    for_to_row,for_to_col=coordinate_finder_insheet(template_transaction,"Payer Group",1)

    Template=border_apply_bottom_u_type(Template,"Transaction",row_no,len(transaction_df)+row_no,col_no,col_no)
    Template=border_apply_bottom_u_type(Template,"Transaction",row_no,len(transaction_df)+row_no,col_no+1,to_col-1)
    Template=border_apply_bottom_u_type(Template,"Transaction",row_no,len(transaction_df)+row_no,to_col,for_to_col)
    
    #Extend Formulas in Sheets
    Template=extend_formulas(Template,"Transaction",transaction_df,for_row_no+2,for_col_no,for_to_col)
    print("Pasting and extension of formulas completed in Transaction Report")
    
    #--------------Deposit Report---------------------
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_deposit,"Facility",1)
    to_row,to_col=coordinate_finder_insheet(template_deposit,"Service Through Date",1)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Deposit",deposit_df,row_no+1,col_no,to_col)
    
    #Formula pasting range finder
    for_row_no,for_col_no=coordinate_finder_insheet(template_deposit,"Months =",1)
    for_to_row,for_to_col=coordinate_finder_insheet(template_deposit,"Payer Group",1)
    
    #Extend Formulas in Sheets
    Template=extend_formulas(Template,"Deposit",deposit_df,for_row_no+2,for_col_no,for_to_col)
    Template=border_apply_bottom_u_type(Template,"Deposit",row_no,len(deposit_df)+row_no,col_no,col_no)
    Template=border_apply_bottom_u_type(Template,"Deposit",row_no,len(deposit_df)+row_no,col_no,to_col)
    
    test_row,test_col=coordinate_finder_insheet(template_deposit,"Test",1)
    Template=border_apply_bottom_u_type(Template,"Deposit",row_no,len(deposit_df)+row_no,for_col_no,test_col-1)
    Template=border_apply_bottom_u_type(Template,"Deposit",row_no,len(deposit_df)+row_no,test_col,for_to_col)
    
    print("Pasting and extension of formulas completed in Deposit Report")
    
    #-----------------DSO------------------
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_dso,"Days of Sales Outstanding",1)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"DSO",dso_df,row_no+2,col_no,col_no+13)
    
    #------Aging-----
    #Payer Grouping copy paste
    row_no, col_no=coordinate_finder_insheet(template_aging,"Types of Payer Group",1)
            
    #Paste values in excel
    Template=copy_paste_raw(Template,"Agings",aging_payers_list,row_no+1,col_no,col_no)
    
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_aging,"Facility Name",1)
    to_row,to_col=coordinate_finder_insheet(template_aging,">365 days",1)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Agings",aging_df,row_no,col_no,to_col)
    
    #Payer Grouping copy paste
    row_no, col_no=coordinate_finder_insheet(template_aging,"Types of Payer Group",1)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Agings",aging_payers_list,row_no+1,col_no,col_no)
    
    print("Pasting in Aging and DSO Completed")
    
    #------Aging Cal-------------
    
    
    #Formula pasting range finder
    for_row_no,for_col_no=coordinate_finder_insheet(border_aging_cal,"% AR over 90",1)
    for_to_row,for_to_col=coordinate_finder_insheet(border_aging_cal,"% AR over 90",first_file+2)
    
    #Extend Formulas in Sheets
    Template=extend_formulas(Template,"Aging Cal",border_aging_cal,for_row_no+2,for_col_no+1,for_col_no+1)
    
    
    #-----------Payer Grouping---------------------------
    
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(template_grouping,"Payer",1)
    row_no_end,col_no_end=coordinate_finder_insheet(template_grouping,"Payer Group",1)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Grouping",payer_group_df,row_no+1,col_no,col_no+1)

    Template=border_apply_bottom_u_type(Template,"Grouping",row_no,len(payer_group_df)+row_no,col_no,col_no_end)
    
    #Payer Grouping copy paste
    row_no, col_no=coordinate_finder_insheet(template_grouping,"Types of Payer Group",1)

    Template=border_apply_bottom_u_type(Template,"Grouping",row_no,no_of_payer_grps+row_no,col_no,col_no)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Grouping",uniques_payer_groups,row_no+1,col_no,col_no)
    
    #----------AR Clean Up Estimates Payers Paste----------------------------
    
    #Getting position
    row_num, col_num=coordinate_finder_insheet(template_clean_up,"Payer Type Summary",1)
    row_end, col_end=coordinate_finder_insheet(template_clean_up,"Payer Type Total",1)
    
    #Paste values in excel
    Template=copy_paste_raw(Template,"Clean Up estimates",wanted_list,row_num+1,col_num,col_num)
    
    row_num, col_num=coordinate_finder_insheet(border_clean_up,"% AR over 90",2)

    Template=extend_formulas(Template,"Clean Up estimates",border_clean_up,row_num+2,col_num+1,col_num+1)
    
    #-------------------DSO & NCR Date range-------------------------
    
    #Find the position of the particular excel paths
    row_no,col_no=coordinate_finder_insheet(border_dsoncr,"Metrics",1)
    
    #To paste the end range month for assessment period
    Template=paste_value_in_cell(Template,"DSO & NCR",row_no,col_no+no_months,end_date)
    
    #--------------------NCR By Payer-------------------------------------
    #Formula pasting range finder
    for_row_no,for_col_no=coordinate_finder_insheet(border_ncr_payer,"Facility",1)
    for_to_row,for_to_col=coordinate_finder_insheet(border_ncr_payer,"Facility",no_of_payer_grps+1)
    
    #Extend Formulas in Sheets
    Template=extend_formulas(Template,"NCR By Payer",border_ncr_payer,for_row_no+2,for_col_no+no_months+2,for_col_no+no_months+2)
    
    #---------------------NCR By Facility---------------------------------
    #Formula pasting range finder
    for_row_no,for_col_no=coordinate_finder_insheet(border_ncr_facility,"Payer Type",1)
    for_to_row,for_to_col=coordinate_finder_insheet(border_ncr_facility,"Next Facility",first_file)
    
    #Extend Formulas in Sheets
    Template=extend_formulas(Template,"NCR By Facility",border_ncr_facility,for_row_no+2,for_col_no+no_months+2,for_col_no+no_months+2)
    print("Hidden column formulas are extended")
    print("Borders updating started....")
    
    # ----------------------Border for NCR Cal-------------------
    
    #Find the position of the particular excel paths iterating over each table and deleting the extra payers here
    
    #The row in which we will be having the border
    interval_border_apply_row=rows_to_keep-1
    
    for i in range(1,first_file+1+1):
        row_no,col_no=coordinate_finder_insheet(border_ncr_cal,"Facility",i)
        
        border_row=row_no+interval_border_apply_row
    
        Template=bottom_border_one_cell(Template,"Payer-Wise NCR Cal",border_row,col_no,col_no)
        
    #-----------NCR By Facility-------------------------
    # for i in range(1,first_file+1+1):
    #     row_no,col_no=coordinate_finder_insheet(template_ncr_facility,"Payer Type",i)
    #     end_row_no,end_col_no=coordinate_finder_insheet(template_ncr_facility,"Next Facility",i)
        
    #     start_row=row_no+no_of_payer_grps-row_reducing+1
    #     no_of_rows_delete=end_row_no-start_row-row_reducing
        
    #     worksheet_ncr_facility=deleting_below_rows(worksheet_ncr_facility,start_row,no_of_rows_delete)
    #     row_reducing+=no_of_rows_delete
        
    #----DSO & NCR DSO Table-----------------
    
    #left border apply 
    row_no,col_no=coordinate_finder_insheet(border_dsoncr,"Facility Name",1)
    start_row=row_no+first_file
    
    Template=left_bottom_border(Template,"DSO & NCR",start_row,col_no,col_no)
    Template=bottom_border(Template,"DSO & NCR",start_row,col_no+1,col_no+14)
    Template=left_right_bottom_border(Template,"DSO & NCR",start_row,col_no+15,col_no+15)
    
    #----Raw Data DSO Tabs--------------
    #left border apply 
    row_no,col_no=coordinate_finder_insheet(border_dso,"Days of Sales Outstanding",1)
    start_row=row_no+2+first_file
    
    Template=left_bottom_border(Template,"DSO",start_row,col_no,col_no)
    Template=bottom_border(Template,"DSO",start_row,col_no+1,col_no+12)
    Template=right_bottom_border(Template,"DSO",start_row,col_no+13,col_no+13)
    
    #Raw Data Aging------------------
    row_no,col_no=coordinate_finder_insheet(border_aging,"Facility Name",1)
    start_row=row_no+len(aging_df)-1
    
    Template=left_bottom_border(Template,"Agings",start_row,col_no,col_no)
    Template=bottom_border(Template,"Agings",start_row,col_no+1,col_no+len(aging_df.columns)-1)
    Template=right_bottom_border(Template,"Agings",start_row,col_no+len(aging_df.columns)-1,col_no+len(aging_df.columns)-1)
    
    #Types of Payers column border
    row_no,col_no=coordinate_finder_insheet(border_aging,"Types of Payer Group",1)
    Template=left_right_bottom_border(Template,"Agings",start_row,col_no,col_no)
    
    #Payer Grouping--------------------
    #Types of Payers column border
    row_no,col_no=coordinate_finder_insheet(border_grouping,"Types of Payer Group",1)
    for i in range(len(uniques_payer_groups)):
        Template=left_right_border(Template,"Grouping",row_no+i,col_no,col_no)
        
    Template=left_right_bottom_border(Template,"Grouping",row_no+len(uniques_payer_groups),col_no,col_no)
    
    row_no,col_no=coordinate_finder_insheet(border_grouping,"Payer",1)
    row_no_end,col_no_end=coordinate_finder_insheet(border_grouping,"Payer Group",1)
    
    Template=border_apply_bottom_u_type(Template,"Grouping",row_no,len(payer_group_df)+row_no,col_no,col_no_end)
    
    print("Borderers updated !!")
    print("Moving the DSO table to the left in DSO & NCR")
    #Moving the right cols of DSo to left
    row_no,col_no=coordinate_finder_insheet(border_dsoncr,"Facility Name",1)
    row_no_dso,col_no_dso=coordinate_finder_insheet(border_dsoncr,"DSO",1)
    row_no_metrics,col_no_metrics=coordinate_finder_insheet(border_dsoncr,"Metrics",1)
    
    
    #geting range to move the cols and rows
    start_row=row_no-2
    start_col=col_no
    end_col=col_no_dso-1
    end_row=row_no+first_file
    
    cell_range = get_cell_range(start_row, start_col, end_row, end_col)
    
    #Cols to move subtracting with metrics to get the approportae range
    columns_to_move=col_no_metrics-1-col_no
    
    Template=cut_cell_range(Template,"DSO & NCR",cell_range,0,columns_to_move)
    
    #Now for the Average alone here we need to translate the avaerage formula 
    
    start_row=row_no_dso-2
    start_col=col_no_dso
    end_col=col_no_dso
    end_row=row_no+first_file
    
    cell_range = get_cell_range(start_row, start_col, end_row, end_col)
    
    Template=cut_cell_range_translate(Template,"DSO & NCR",cell_range,0,columns_to_move)
    print("Assessment completed, saving the file.....")
    
    Template.save(file_saving_location)
    perform_calc_delete_extra_sheet(file_saving_location,"Evaluation Warning")
    
    print("Assessment completed and saved!!!")
    #Timer
    end_time = time.time()
    elapsed_time_seconds = end_time - start_time
    elapsed_minutes = int(elapsed_time_seconds // 60)
    elapsed_seconds = int(elapsed_time_seconds % 60)
    print("Elapsed time: ", elapsed_minutes, "minutes", elapsed_seconds, "seconds")

    #Notification
    from IPython.display import Audio
    # Play the audio
    Audio(filename=audio_file_path, autoplay=True)
    
    # Payers Extraction
    T_df=create_df_from_sheet("Transaction",file_saving_location)
    D_df=create_df_from_sheet("Deposit",file_saving_location)
    
    T_df.columns=T_df.iloc[0]
    T_df=T_df.iloc[1:,:].reset_index(drop=True)
    cutoff_date = pd.Timestamp('1900-01-01')
    T_df['SOM Service through date'] = pd.to_datetime(T_df['SOM Service through date'], errors='coerce')
    # Filter rows where 'SOM Service through date' is greater than the cutoff date
    T_df = T_df[T_df['SOM Service through date'] > cutoff_date]
    
    #Find the position of the particular excel paths
    search_row,search_col=coordinate_finder_insheet(template_transaction,"Test",1)
    value_row,value_col=coordinate_finder_insheet(template_transaction,"Payer",1)
    
    T_payers_df=get_row_with_condition(T_df,search_col-1,value_col-1,1)
                
    #Find the position of the particular excel paths
    search_row,search_col=coordinate_finder_insheet(template_deposit,"Test",1)
    value_row,value_col=coordinate_finder_insheet(template_deposit,"Private Payer Check",1)
    
    #Getting payers from the dataframe
    D_payers_df=get_row_with_condition(D_df,search_col-1,value_col,1)
    
    payers_df=concatenate_two_df(T_payers_df,D_payers_df)
    
    payers_df=unique_row_values(payers_df,0)
    
    
    #Pasting the raw data in the Payer group template
    payer_grp_file=load_workbook(payer_grp_send)
    
    #Find the position of the particular excel paths
    payer_grp_send_df=pd.read_excel(payer_grp_send,header=None)
    row_no,col_no=coordinate_finder_insheet(payer_grp_send_df,"Payer",1)
    
    #Paste values in excel
    payer_grp_file=copy_paste_raw(payer_grp_file,"Payer Grouping",payers_df,row_no+1,col_no,col_no)
    
    start_row=row_no+len(payers_df)
    payer_grp_file=left_bottom_border(payer_grp_file,"Payer Grouping",start_row,col_no,col_no)
    
    payer_grp_file=right_bottom_border(payer_grp_file,"Payer Grouping",start_row,col_no+1,col_no+1)
    
    #Saving the file
    payer_grp_file.save(payer_grp_mail)
    
    workbook2 = Workbook()
    workbook2.LoadFromFile(payer_grp_mail)
    
    worksheet_below_row = workbook2.Worksheets["Payer Grouping"] 
    worksheet_below_row = deleting_below_rows(worksheet_below_row,start_row+1,10000)
    
    workbook2.SaveToFile(payer_grp_mail)
    
    print("Payer Group file is Generated!")
    
    

